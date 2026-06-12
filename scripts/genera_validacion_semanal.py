"""Genera reporte HTML de validacion semanal: Realidad vs Pronostico.

Compara los casos reales del boletin epidemiologico SINAVE con las predicciones
de los modelos de produccion para la semana mas reciente disponible.
Incluye desglose por sexo (Nacional y Regional) y actualiza la columna
realidad_sem_previa en la tabla de produccion Excel.

Salida: reports/ProdDetails/validacion_semanal.html
"""

from __future__ import annotations

from datetime import date, datetime
import math
from pathlib import Path
from typing import Any
from zoneinfo import ZoneInfo

import pandas as pd

from epiforecast.utils.cohorts import filter_neuro

# ---------------------------------------------------------------------------
# Rutas y constantes
# ---------------------------------------------------------------------------
ROOT = Path(__file__).resolve().parent.parent
BOLETIN = ROOT / "data" / "processed" / "dataset_boletin_epidemiologico.csv"
FORECASTS_DIR = ROOT / "reports" / "forecasts"
PROD_EXCEL = ROOT / "reports" / "ProdDetails" / "tabla_333_modelos_produccion.xlsx"
OUTPUT = ROOT / "reports" / "ProdDetails" / "validacion_semanal.html"

MODELS = ["prophet", "deepar", "ensemble", "stacking"]
PADECIMIENTOS = ["Alzheimer", "Parkinson", "Depresion"]
PAD_DISPLAY: dict[str, str] = {
    "Alzheimer": "Alzheimer (G30)",
    "Parkinson": "Parkinson (G20)",
    "Depresion": "Depresion (F32)",
}
SEXO_MODES = ["general", "hombres", "mujeres"]
SEXO_DISPLAY: dict[str, str] = {
    "general": "General",
    "hombres": "Hombres",
    "mujeres": "Mujeres",
}

ENTITY_ORDER = [
    "Nacional",
    "Aguascalientes",
    "Baja California",
    "Baja California Sur",
    "Campeche",
    "Chiapas",
    "Chihuahua",
    "Ciudad de Mexico",
    "Coahuila",
    "Colima",
    "Durango",
    "Guanajuato",
    "Guerrero",
    "Hidalgo",
    "Jalisco",
    "Michoacan",
    "Mexico",
    "Morelos",
    "Nayarit",
    "Nuevo Leon",
    "Oaxaca",
    "Puebla",
    "Queretaro",
    "Quintana Roo",
    "San Luis Potosi",
    "Sinaloa",
    "Sonora",
    "Tabasco",
    "Tamaulipas",
    "Tlaxcala",
    "Veracruz",
    "Yucatan",
    "Zacatecas",
]

REGION_STATES: dict[str, list[str]] = {
    "Region Metropolitana alta": [
        "Ciudad de Mexico",
        "Mexico",
        "Nuevo Leon",
        "Jalisco",
    ],
    "Region Urbana media": [
        "Aguascalientes",
        "Baja California",
        "Baja California Sur",
        "Chihuahua",
        "Coahuila",
        "Colima",
        "Durango",
        "Guanajuato",
        "Morelos",
        "Queretaro",
        "San Luis Potosi",
        "Sinaloa",
        "Sonora",
        "Tamaulipas",
        "Zacatecas",
    ],
    "Region Rural / dispersa": [
        "Guerrero",
        "Hidalgo",
        "Michoacan",
        "Nayarit",
        "Puebla",
        "Tlaxcala",
        "Veracruz",
    ],
    "Region Sur-Sureste vulnerable": [
        "Campeche",
        "Chiapas",
        "Oaxaca",
        "Tabasco",
        "Yucatan",
        "Quintana Roo",
    ],
}
REGION_ORDER = list(REGION_STATES.keys())
REGION_DISPLAY: dict[str, str] = {
    "Region Metropolitana alta": "Metropolitana alta",
    "Region Urbana media": "Urbana media",
    "Region Rural / dispersa": "Rural / dispersa",
    "Region Sur-Sureste vulnerable": "Sur-Sureste vulnerable",
}
# Reverse lookup: state -> region
_STATE_TO_REGION: dict[str, str] = {}
for _reg, _states in REGION_STATES.items():
    for _st in _states:
        _STATE_TO_REGION[_st] = _reg

TZ_MX = ZoneInfo("America/Mexico_City")

# ---------------------------------------------------------------------------
# Normalizacion
# ---------------------------------------------------------------------------


def _normalize(s: str) -> str:
    """Quita acentos para hacer join entre fuentes."""
    import unicodedata

    nfkd = unicodedata.normalize("NFKD", s)
    return "".join(c for c in nfkd if not unicodedata.combining(c))


def _norm_prod_entidad(s: str) -> str:
    """Normaliza entidad del Excel de produccion (region_ -> Region )."""
    if s.startswith("region_"):
        s = "Region " + s[len("region_") :]
    return _normalize(s)


# ---------------------------------------------------------------------------
# Carga de datos
# ---------------------------------------------------------------------------


def _epiweek_to_monday(anio: int, semana: int) -> date:
    """Convierte semana epidemiologica a fecha del lunes (ISO)."""
    return date.fromisocalendar(anio, semana, 1)


def _epiweek_date_range(anio: int, semana: int) -> str:
    """Rango de fechas legible."""
    monday = _epiweek_to_monday(anio, semana)
    sunday = date.fromisocalendar(anio, semana, 7)
    meses = [
        "",
        "enero",
        "febrero",
        "marzo",
        "abril",
        "mayo",
        "junio",
        "julio",
        "agosto",
        "septiembre",
        "octubre",
        "noviembre",
        "diciembre",
    ]
    if monday.month == sunday.month:
        return f"{monday.day}-{sunday.day} {meses[monday.month]} {monday.year}"
    return f"{monday.day} {meses[monday.month]} - {sunday.day} {meses[sunday.month]} {sunday.year}"


def _load_boletin_full() -> tuple[pd.DataFrame, int, int]:
    """Carga boletin con desglose por sexo derivado de acumulados.

    Returns (df, anio, semana) donde df tiene columnas:
      entidad_norm, pad_norm, sexo, real
    Solo filas a nivel estado (32 entidades).
    """
    df = pd.read_csv(BOLETIN)
    # Guard: la validación semanal es solo de la cohorte neuro de producción.
    # Excluye Dengue (en el consolidado pero sin forecasts) para no generar filas fantasma.
    df = filter_neuro(df)
    anio = int(df["Anio"].max())
    semana = int(df.loc[df["Anio"] == anio, "Semana"].max())

    df_cur = df[(df["Anio"] == anio) & (df["Semana"] == semana)].copy()

    # Semana previa para derivar hombres/mujeres
    if semana > 1:
        df_prev = df[(df["Anio"] == anio) & (df["Semana"] == semana - 1)].copy()
    else:
        prev_anio = anio - 1
        prev_sem = int(df.loc[df["Anio"] == prev_anio, "Semana"].max())
        df_prev = df[(df["Anio"] == prev_anio) & (df["Semana"] == prev_sem)].copy()

    df_cur["entidad_norm"] = df_cur["Entidad"].apply(_normalize)
    df_cur["pad_norm"] = df_cur["Padecimiento"].apply(_normalize)
    df_prev["entidad_norm"] = df_prev["Entidad"].apply(_normalize)
    df_prev["pad_norm"] = df_prev["Padecimiento"].apply(_normalize)

    rows: list[dict[str, Any]] = []
    for _, cur in df_cur.iterrows():
        ent = cur["entidad_norm"]
        pad = cur["pad_norm"]

        # General
        rows.append(
            {
                "entidad_norm": ent,
                "pad_norm": pad,
                "sexo": "general",
                "real": cur["Casos_semana"],
            }
        )

        # Hombres / mujeres desde diferencia de acumulados
        prev = df_prev[(df_prev["entidad_norm"] == ent) & (df_prev["pad_norm"] == pad)]
        h_week: float | None = None
        m_week: float | None = None
        if not prev.empty:
            h_cur = cur["Acumulado_hombres"]
            h_prev = prev.iloc[0]["Acumulado_hombres"]
            m_cur = cur["Acumulado_mujeres"]
            m_prev = prev.iloc[0]["Acumulado_mujeres"]
            if pd.notna(h_cur) and pd.notna(h_prev):
                h_week = max(0, h_cur - h_prev)
            if pd.notna(m_cur) and pd.notna(m_prev):
                m_week = max(0, m_cur - m_prev)

        rows.append(
            {
                "entidad_norm": ent,
                "pad_norm": pad,
                "sexo": "hombres",
                "real": h_week,
            }
        )
        rows.append(
            {
                "entidad_norm": ent,
                "pad_norm": pad,
                "sexo": "mujeres",
                "real": m_week,
            }
        )

    return pd.DataFrame(rows), anio, semana


def _load_forecasts(target_date: str) -> pd.DataFrame:
    """Lee los 4 all_forecast CSVs, filtra por fecha. Incluye todos los modos."""
    frames = []
    for model in MODELS:
        path = FORECASTS_DIR / model / f"all_forecast_{model}.csv"
        if not path.exists():
            continue
        df = pd.read_csv(path)
        df["ds"] = pd.to_datetime(df["ds"]).dt.strftime("%Y-%m-%d")
        df = df[df["ds"] == target_date].copy()
        df["model"] = model
        df["entidad_norm"] = df["meta_entidad"].apply(_normalize)
        df["pad_norm"] = df["meta_padecimiento"].apply(_normalize)
        df.rename(columns={"meta_modo": "sexo"}, inplace=True)
        frames.append(df[["entidad_norm", "pad_norm", "sexo", "yhat", "model"]])
    if not frames:
        msg = f"No se encontraron forecasts para {target_date}"
        raise FileNotFoundError(msg)
    return pd.concat(frames, ignore_index=True)


def _load_production() -> pd.DataFrame:
    """Lee Excel de produccion, retorna motor ganador y MASE para todas las combinaciones."""
    df = pd.read_excel(PROD_EXCEL, sheet_name=0)
    df = df[["padecimiento", "entidad", "sexo", "modelo_produccion", "mase_prod"]].copy()
    df["entidad_norm"] = df["entidad"].apply(_norm_prod_entidad)
    df["pad_norm"] = df["padecimiento"].apply(_normalize)
    df["motor"] = df["modelo_produccion"]
    return df[["entidad_norm", "pad_norm", "sexo", "motor", "mase_prod"]]


# ---------------------------------------------------------------------------
# Calculos
# ---------------------------------------------------------------------------


def _verdict(real_val: Any, yhat: Any) -> tuple[float | None, str]:
    """Calcula SMAPE y veredicto.

    SMAPE = |F - A| / ((|A| + |F|) / 2) * 100
    Rango: 0-200%. Simetrico, no explota con incidencia baja.
    """
    if real_val is None or yhat is None:
        return None, "N/A"
    if isinstance(real_val, float) and math.isnan(real_val):
        return None, "N/A"
    if isinstance(yhat, float) and math.isnan(yhat):
        return None, "N/A"
    real_val = round(real_val)
    yhat = round(yhat)
    denom = (abs(real_val) + abs(yhat)) / 2
    if denom == 0:
        # Ambos cero: pronostico perfecto
        return 0.0, "EXCELENTE"
    if real_val == 0 and yhat <= 2:
        # Cero real con pronostico despreciable
        return None, "N/A"
    smape = abs(yhat - real_val) / denom * 100
    if smape <= 10:
        return smape, "EXCELENTE"
    if smape <= 25:
        return smape, "BUENO"
    if smape <= 50:
        return smape, "REGULAR"
    return smape, "FALLO"


def _build_comparison(
    boletin: pd.DataFrame,
    forecasts: pd.DataFrame,
    produccion: pd.DataFrame,
) -> pd.DataFrame:
    """Construye tabla de comparacion con ~333 filas.

    Niveles:
      - nacional: 3 pads x 3 sexos = 9
      - regional: 4 regiones x 3 pads x 3 sexos = 36
      - estatal:  32 estados x 3 pads x 3 sexos = 288
    Total: 333
    """
    all_rows: list[dict[str, Any]] = []

    # --- Nacional: sumar estados por (pad, sexo) ---
    for pad in boletin["pad_norm"].unique():
        for sexo in SEXO_MODES:
            mask = (boletin["pad_norm"] == pad) & (boletin["sexo"] == sexo)
            total = boletin.loc[mask, "real"].sum()
            all_rows.append(
                {
                    "entidad_norm": "Nacional",
                    "pad_norm": pad,
                    "sexo": sexo,
                    "real": total,
                    "nivel": "nacional",
                }
            )

    # --- Regional: sumar estados por region ---
    for region, states in REGION_STATES.items():
        for pad in boletin["pad_norm"].unique():
            for sexo in SEXO_MODES:
                mask = (
                    boletin["entidad_norm"].isin(states)
                    & (boletin["pad_norm"] == pad)
                    & (boletin["sexo"] == sexo)
                )
                total = boletin.loc[mask, "real"].sum()
                all_rows.append(
                    {
                        "entidad_norm": region,
                        "pad_norm": pad,
                        "sexo": sexo,
                        "real": total,
                        "nivel": "regional",
                    }
                )

    # --- Estatal: filas individuales ---
    for _, row in boletin.iterrows():
        all_rows.append(
            {
                "entidad_norm": row["entidad_norm"],
                "pad_norm": row["pad_norm"],
                "sexo": row["sexo"],
                "real": row["real"],
                "nivel": "estatal",
            }
        )

    real_df = pd.DataFrame(all_rows)

    # Merge con produccion para motor ganador
    merged = real_df.merge(
        produccion,
        on=["entidad_norm", "pad_norm", "sexo"],
        how="left",
    )

    # Lookup forecast yhat del motor ganador
    results: list[dict[str, Any]] = []
    for _, row in merged.iterrows():
        motor = row.get("motor")
        if pd.isna(motor) or not motor:
            motor = "ensemble"
        motor_lower = str(motor).lower()

        match = forecasts[
            (forecasts["entidad_norm"] == row["entidad_norm"])
            & (forecasts["pad_norm"] == row["pad_norm"])
            & (forecasts["sexo"] == row["sexo"])
            & (forecasts["model"] == motor_lower)
        ]
        yhat = round(match["yhat"].iloc[0]) if len(match) > 0 else None

        real_val = row["real"]
        if real_val is not None and not (isinstance(real_val, float) and math.isnan(real_val)):
            real_val = round(real_val)

        error_pct, veredicto = _verdict(real_val, yhat)

        mase_val = row.get("mase_prod")
        mase_val = mase_val if pd.notna(mase_val) else None

        results.append(
            {
                "entidad_norm": row["entidad_norm"],
                "pad_norm": row["pad_norm"],
                "sexo": row["sexo"],
                "real": real_val,
                "pronostico": yhat,
                "error_pct": error_pct,
                "veredicto": veredicto,
                "motor": motor_lower,
                "mase": mase_val,
                "nivel": row["nivel"],
            }
        )

    return pd.DataFrame(results)


# ---------------------------------------------------------------------------
# HTML helpers
# ---------------------------------------------------------------------------

_FONTS_LINK = (
    '<link rel="preconnect" href="https://fonts.googleapis.com">'
    '<link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>'
    '<link href="https://fonts.googleapis.com/css2?family=DM+Serif+Display'
    "&family=JetBrains+Mono:wght@400;600"
    "&family=Source+Sans+3:wght@300;400;600;700"
    '&display=swap" rel="stylesheet">'
)

VEREDICTO_COLORS: dict[str, str] = {
    "EXCELENTE": "#2DD4BF",
    "BUENO": "#0D9488",
    "REGULAR": "#F59E0B",
    "FALLO": "#EF4444",
    "N/A": "#97999B",
}
VEREDICTO_TEXT_COLOR: dict[str, str] = {
    "EXCELENTE": "#0B0F1A",
    "BUENO": "#fff",
    "REGULAR": "#0B0F1A",
    "FALLO": "#fff",
    "N/A": "#fff",
}
MOTOR_COLORS: dict[str, str] = {
    "prophet": "#5B8DEF",
    "deepar": "#BE185D",
    "ensemble": "#FF6F00",
    "stacking": "#1A237E",
}


def _badge(text: str, bg: str, fg: str = "#fff") -> str:
    return (
        f'<span style="display:inline-flex;align-items:center;padding:.25rem .85rem;'
        f"border-radius:100px;font-size:.78rem;font-weight:700;"
        f'color:{fg};background:{bg}">{text}</span>'
    )


def _veredicto_badge(v: str) -> str:
    return _badge(v, VEREDICTO_COLORS.get(v, "#97999B"), VEREDICTO_TEXT_COLOR.get(v, "#fff"))


def _motor_badge(m: str) -> str:
    return _badge(m.capitalize(), MOTOR_COLORS.get(m, "#97999B"))


def _semaforo_dot(v: str) -> str:
    color = VEREDICTO_COLORS.get(v, "#97999B")
    return (
        f'<span title="{v}" style="display:inline-block;width:22px;height:22px;'
        f"border-radius:50%;background:{color};border:2px solid #fff;"
        f'box-shadow:0 1px 4px rgba(0,0,0,.15)"></span>'
    )


def _fmt_error(val: float | None) -> str:
    if val is None or (isinstance(val, float) and math.isnan(val)):
        return "—"
    return f"{val:.1f}%"


def _fmt_int(val: Any) -> str:
    if val is None or (isinstance(val, float) and math.isnan(val)):
        return "N/A"
    return f"{int(val):,}"


def _err_class(error_pct: float | None) -> str:
    if error_pct is None or (isinstance(error_pct, float) and math.isnan(error_pct)):
        return ""
    if error_pct <= 25:
        return "err-good"
    return "err-bad"


def _fmt_mase(val: Any) -> str:
    if val is None or (isinstance(val, float) and math.isnan(val)):
        return "N/A"
    return f"{val:.2f}"


def _mase_class(val: Any) -> str:
    if val is None or (isinstance(val, float) and math.isnan(val)):
        return ""
    if val < 1.0:
        return "mase-good"
    return "mase-warn"


MASE_VERDICT_COLORS: dict[str, str] = {
    "EXCELENTE": "#2DD4BF",
    "BUENO": "#0D9488",
    "REGULAR": "#F59E0B",
    "FALLO": "#EF4444",
    "—": "#97999B",
}
MASE_VERDICT_TEXT: dict[str, str] = {
    "EXCELENTE": "#0B0F1A",
    "BUENO": "#fff",
    "REGULAR": "#0B0F1A",
    "FALLO": "#fff",
    "—": "#fff",
}


def _mase_verdict(val: Any) -> str:
    """Veredicto basado en MASE: < 1 supera al naive."""
    if val is None or (isinstance(val, float) and math.isnan(val)):
        return "—"
    if val < 0.5:
        return "EXCELENTE"
    if val < 1.0:
        return "BUENO"
    if val < 1.5:
        return "REGULAR"
    return "FALLO"


def _mase_dot(val: Any) -> str:
    """Dot de color segun veredicto MASE, sin etiqueta."""
    v = _mase_verdict(val)
    color = MASE_VERDICT_COLORS.get(v, "#97999B")
    return (
        f'<span title="MASE {v}" style="display:inline-block;width:10px;height:10px;'
        f"border-radius:50%;background:{color};vertical-align:middle;"
        f'margin-left:.4rem"></span>'
    )


# ---------------------------------------------------------------------------
# Tabla reutilizable con desglose por sexo
# ---------------------------------------------------------------------------


def _build_sexo_table(data: pd.DataFrame) -> str:
    """Tabla con filas agrupadas: padecimiento (rowspan=3) x sexo."""
    rows_html = ""
    for pad in PADECIMIENTOS:
        pad_rows = data[data["pad_norm"] == pad]
        if pad_rows.empty:
            continue
        n_sexos = sum(1 for s in SEXO_MODES if not pad_rows[pad_rows["sexo"] == s].empty)
        first = True
        for sexo in SEXO_MODES:
            sr = pad_rows[pad_rows["sexo"] == sexo]
            if sr.empty:
                continue
            r = sr.iloc[0]
            is_last = sexo == SEXO_MODES[-1]
            row_cls = ' class="row-group-last"' if is_last else ""
            pad_cell = ""
            if first:
                pad_cell = (
                    f'<td rowspan="{n_sexos}" style="font-family:var(--font-body);'
                    f'font-weight:600;vertical-align:middle">'
                    f"{PAD_DISPLAY.get(pad, pad)}</td>"
                )
                first = False
            rows_html += (
                f"<tr{row_cls}>"
                f"  {pad_cell}"
                f'  <td style="font-family:var(--font-body)">{SEXO_DISPLAY.get(sexo, sexo)}</td>'
                f"  <td>{_fmt_int(r['real'])}</td>"
                f"  <td>{_fmt_int(r['pronostico'])}</td>"
                f'  <td class="{_err_class(r["error_pct"])}">'
                f"{_fmt_error(r['error_pct'])}</td>"
                f'  <td class="{_mase_class(r["mase"])}">'
                f"{_fmt_mase(r['mase'])} {_mase_dot(r['mase'])}</td>"
                f"  <td>{_motor_badge(r['motor'])}</td>"
                f"  <td>{_veredicto_badge(r['veredicto'])}</td>"
                f"</tr>\n"
            )

    return f"""<div class="table-wrapper">
  <table>
    <thead><tr>
      <th>Padecimiento</th><th>Sexo</th><th>Real</th><th>Pronostico</th>
      <th>SMAPE</th><th>MASE</th><th>Motor</th><th>Veredicto</th>
    </tr></thead>
    <tbody>{rows_html}</tbody>
  </table>
</div>"""


# ---------------------------------------------------------------------------
# Secciones HTML
# ---------------------------------------------------------------------------


def _section_dashboard(comp: pd.DataFrame, anio: int, semana: int) -> str:
    """Panel premium con gauges SVG animados: SMAPE + MASE por padecimiento."""
    nac_gen = comp[(comp["nivel"] == "nacional") & (comp["sexo"] == "general")]

    # MASE desde tabla de produccion (validacion cruzada del modelo ganador)
    prod_df = pd.read_excel(PROD_EXCEL, sheet_name=0)
    nac_prod = prod_df[(prod_df["entidad"] == "Nacional") & (prod_df["sexo"] == "general")]
    mase_lookup: dict[str, float] = {}
    for _, row in nac_prod.iterrows():
        mase_lookup[_normalize(row["padecimiento"])] = row.get("mase_prod", float("nan"))

    date_range = _epiweek_date_range(anio, semana)
    panels = ""
    for pad in PADECIMIENTOS:
        row = nac_gen[nac_gen["pad_norm"] == pad]
        if row.empty:
            continue
        r = row.iloc[0]
        smape = r["error_pct"] if pd.notna(r["error_pct"]) else 0.0
        real = int(r["real"]) if pd.notna(r["real"]) else 0
        pron = int(r["pronostico"]) if pd.notna(r["pronostico"]) else 0
        motor = r["motor"]
        veredicto = r["veredicto"]
        mase = mase_lookup.get(pad, float("nan"))
        panels += _dashboard_panel(pad, smape, real, pron, motor, veredicto, mase)

    return f"""
<section class="reveal animate">
  <div class="dashboard">
    <div class="dashboard-header">
      <h2>Panel de Desempeno Nacional</h2>
      <p>Semana {semana} | {date_range} | Modo general</p>
    </div>
    <div class="dashboard-grid">{panels}</div>
  </div>
</section>"""


# Gauge constants
_GAUGE_SIZE = 160
_GAUGE_R = 65
_GAUGE_CIRC = 2 * math.pi * _GAUGE_R  # 408.41


def _dashboard_panel(
    pad: str,
    smape: float,
    real: int,
    pron: int,
    motor: str,
    veredicto: str,
    mase: float,
) -> str:
    """Panel individual de un padecimiento con gauge SVG + MASE bar."""
    cx = cy = _GAUGE_SIZE / 2
    precision = max(0.0, min(100.0, 100.0 - smape))
    target_offset = _GAUGE_CIRC * (1 - precision / 100)
    gauge_color = VEREDICTO_COLORS.get(veredicto, "#97999B")

    # MASE bar
    mase_html = ""
    if not math.isnan(mase):
        mase_pct = min(mase / 2 * 100, 100)
        if mase < 0.8:
            mase_color, mase_lbl = "#2DD4BF", "Excelente"
        elif mase < 1.0:
            mase_color, mase_lbl = "#0D9488", "Superior al naive"
        elif mase < 1.3:
            mase_color, mase_lbl = "#F59E0B", "Comparable al naive"
        else:
            mase_color, mase_lbl = "#EF4444", "Inferior al naive"
        mase_html = f"""
      <div class="mase-box">
        <div class="mase-head">
          <span class="mase-title">MASE</span>
          <span class="mase-val" style="color:{mase_color}">{mase:.2f}</span>
        </div>
        <div class="mase-track">
          <div class="mase-fill" style="--mw:{mase_pct}%;background:{mase_color}"></div>
          <div class="mase-ref"></div>
        </div>
        <div class="mase-ticks">
          <span>0</span>
          <span style="position:absolute;left:50%;transform:translateX(-50%)">1.0</span>
          <span>2.0</span>
        </div>
        <div class="mase-lbl" style="color:{mase_color}">{mase_lbl}</div>
      </div>"""

    # Flecha SVG entre Real y Pronostico
    arrow = (
        '<svg width="28" height="28" viewBox="0 0 24 24" fill="none" '
        'stroke="rgba(232,213,181,.35)" stroke-width="2" stroke-linecap="round">'
        '<path d="M5 12h14M13 6l6 6-6 6"/></svg>'
    )

    return f"""
    <div class="dash-panel">
      <h3>{PAD_DISPLAY.get(pad, pad)}</h3>
      <div class="gauge-wrap">
        <svg width="{_GAUGE_SIZE}" height="{_GAUGE_SIZE}"
             viewBox="0 0 {_GAUGE_SIZE} {_GAUGE_SIZE}">
          <defs>
            <filter id="glow-{pad}" x="-50%" y="-50%" width="200%" height="200%">
              <feGaussianBlur stdDeviation="4" result="blur"/>
              <feMerge><feMergeNode in="blur"/><feMergeNode in="SourceGraphic"/></feMerge>
            </filter>
          </defs>
          <circle cx="{cx}" cy="{cy}" r="{_GAUGE_R}"
            fill="none" stroke="rgba(255,255,255,.07)" stroke-width="11"/>
          <circle cx="{cx}" cy="{cy}" r="{_GAUGE_R}"
            fill="none" stroke="{gauge_color}" stroke-width="11"
            stroke-dasharray="{_GAUGE_CIRC:.2f}"
            stroke-linecap="round"
            transform="rotate(-90 {cx} {cy})"
            filter="url(#glow-{pad})"
            class="gauge-arc"
            style="--to:{target_offset:.2f}"/>
          <text x="{cx}" y="{cy - 10}" text-anchor="middle"
            fill="#fff" style="font-family:var(--font-display);font-size:28px">
            {smape:.1f}%</text>
          <text x="{cx}" y="{cy + 14}" text-anchor="middle"
            fill="rgba(232,213,181,.55)" style="font-family:var(--font-body);font-size:12px">
            SMAPE</text>
        </svg>
      </div>
      <div class="dash-nums">
        <div class="dash-n">
          <span class="dash-v" data-target="{real}">{real:,}</span>
          <span class="dash-l">Real</span>
        </div>
        {arrow}
        <div class="dash-n">
          <span class="dash-v" data-target="{pron}">{pron:,}</span>
          <span class="dash-l">Pronostico</span>
        </div>
      </div>
      {mase_html}
      <div class="dash-badges">
        {_motor_badge(motor)} {_veredicto_badge(veredicto)}
      </div>
    </div>"""


def _section_nacional(comp: pd.DataFrame) -> str:
    nac = comp[comp["nivel"] == "nacional"]
    table = _build_sexo_table(nac)
    return f"""
<section class="reveal animate">
  <h2 class="section-title">Resumen Nacional</h2>
  <p class="section-sub">Comparacion agregada por padecimiento y sexo</p>
  <div class="card card-gold">{table}</div>
</section>"""


def _section_regional(comp: pd.DataFrame) -> str:
    reg = comp[comp["nivel"] == "regional"]
    parts = [
        """
<section class="reveal animate">
  <h2 class="section-title">Desempeno Regional</h2>
  <p class="section-sub">4 regiones sociales con desglose por sexo</p>"""
    ]
    for region in REGION_ORDER:
        rd = reg[reg["entidad_norm"] == region]
        if rd.empty:
            continue
        display = REGION_DISPLAY.get(region, region)
        n_states = len(REGION_STATES.get(region, []))
        table = _build_sexo_table(rd)
        parts.append(
            f"""
  <div class="card reveal animate">
    <h2>{display} <span style="font-size:.85rem;color:var(--cool-gray);
      font-family:var(--font-body);font-weight:400">({n_states} entidades)</span></h2>
    {table}
  </div>"""
        )
    parts.append("</section>")
    return "".join(parts)


def _section_detalle(comp: pd.DataFrame) -> str:
    """Detalle por padecimiento: solo modo general, 33 geos."""
    general = comp[(comp["nivel"] == "estatal") & (comp["sexo"] == "general")]
    parts = [
        """
<section class="reveal animate">
  <h2 class="section-title">Detalle por Entidad Federativa</h2>
  <p class="section-sub">Modo general &mdash; 32 entidades + Nacional</p>"""
    ]

    # Agregar Nacional general al inicio
    nac_general = comp[(comp["nivel"] == "nacional") & (comp["sexo"] == "general")]
    combined = pd.concat([nac_general, general], ignore_index=True)

    for pad in PADECIMIENTOS:
        pad_data = combined[combined["pad_norm"] == pad].copy()
        if pad_data.empty:
            continue

        order_map = {e: i for i, e in enumerate(ENTITY_ORDER)}
        pad_data = pad_data.assign(
            _order=pad_data["entidad_norm"].map(lambda x, om=order_map: om.get(x, 999))
        )
        pad_data = pad_data.sort_values("_order")

        rows_html = ""
        total_real = 0.0
        total_prod = 0.0
        for _, r in pad_data.iterrows():
            is_nac = r["entidad_norm"] == "Nacional"
            row_class = ' class="row-nacional"' if is_nac else ""
            real_v = r["real"] if r["real"] is not None and not _isnan(r["real"]) else 0
            prod_v = (
                r["pronostico"]
                if r["pronostico"] is not None and not _isnan(r["pronostico"])
                else 0
            )
            if not is_nac:
                total_real += real_v
                total_prod += prod_v
            rows_html += (
                f"<tr{row_class}>"
                f'  <td style="font-family:var(--font-body)">{r["entidad_norm"]}</td>'
                f"  <td>{_fmt_int(r['real'])}</td>"
                f"  <td>{_fmt_int(r['pronostico'])}</td>"
                f'  <td class="{_err_class(r["error_pct"])}">'
                f"{_fmt_error(r['error_pct'])}</td>"
                f'  <td class="{_mase_class(r["mase"])}">'
                f"{_fmt_mase(r['mase'])} {_mase_dot(r['mase'])}</td>"
                f"  <td>{_motor_badge(r['motor'])}</td>"
                f"  <td>{_veredicto_badge(r['veredicto'])}</td>"
                f"</tr>\n"
            )

        # Conteo de veredictos
        n_exc = len(pad_data[pad_data["veredicto"] == "EXCELENTE"])
        n_good = len(pad_data[pad_data["veredicto"] == "BUENO"])
        n_reg = len(pad_data[pad_data["veredicto"] == "REGULAR"])
        n_fail = len(pad_data[pad_data["veredicto"] == "FALLO"])
        n_na = len(pad_data[pad_data["veredicto"] == "N/A"])

        precision_total = ""
        if total_real > 0:
            prec = abs(total_prod - total_real) / total_real * 100
            precision_total = f" | Error acumulado: {prec:.1f}%"

        parts.append(f"""
  <div class="card reveal animate">
    <h2>{PAD_DISPLAY.get(pad, pad)}</h2>
    <div class="table-wrapper">
      <table><thead><tr>
        <th>Entidad</th><th>Real</th><th>Pronostico</th>
        <th>SMAPE</th><th>MASE</th><th>Motor</th><th>Veredicto</th>
      </tr></thead>
      <tbody>{rows_html}</tbody></table>
    </div>
    <div class="card-footer">
      <span>{_veredicto_badge("EXCELENTE")} {n_exc}</span>
      <span>{_veredicto_badge("BUENO")} {n_good}</span>
      <span>{_veredicto_badge("REGULAR")} {n_reg}</span>
      <span>{_veredicto_badge("FALLO")} {n_fail}</span>
      <span>{_veredicto_badge("N/A")} {n_na}</span>
      <span style="margin-left:auto;font-size:.85rem;color:var(--cool-gray)">
        Total: {_fmt_int(total_real)} real / {_fmt_int(total_prod)} pron{precision_total}
      </span>
    </div>
  </div>""")

    parts.append("</section>")
    return "".join(parts)


def _section_semaforo(comp: pd.DataFrame) -> str:
    """Grid 32x3 solo modo general."""
    general = comp[(comp["nivel"] == "estatal") & (comp["sexo"] == "general")]
    states = [e for e in ENTITY_ORDER if e != "Nacional"]
    header = "<tr><th>Entidad</th>"
    for pad in PADECIMIENTOS:
        header += f"<th>{PAD_DISPLAY.get(pad, pad)}</th>"
    header += "</tr>"

    rows = ""
    for state in states:
        rows += f'<tr><td style="font-family:var(--font-body);font-weight:500">{state}</td>'
        for pad in PADECIMIENTOS:
            match = general[(general["entidad_norm"] == state) & (general["pad_norm"] == pad)]
            if match.empty:
                rows += f"<td>{_semaforo_dot('N/A')}</td>"
            else:
                v = match.iloc[0]["veredicto"]
                err_str = _fmt_error(match.iloc[0]["error_pct"])
                rows += (
                    f'<td style="text-align:center">{_semaforo_dot(v)}'
                    f'<br><span style="font-size:.7rem;color:var(--cool-gray)">'
                    f"{err_str}</span></td>"
                )
        rows += "</tr>"

    leyenda = '<div class="leyenda">'
    for label, color in VEREDICTO_COLORS.items():
        desc = {
            "EXCELENTE": "SMAPE &le;10%",
            "BUENO": "SMAPE &le;25%",
            "REGULAR": "SMAPE &le;50%",
            "FALLO": "SMAPE &gt;50%",
            "N/A": "Sin datos / cero real",
        }[label]
        leyenda += (
            f'<span class="leyenda-item">'
            f'<span style="display:inline-block;width:14px;height:14px;border-radius:50%;'
            f'background:{color};vertical-align:middle;margin-right:.3rem"></span>'
            f"{label}: {desc}</span>"
        )
    leyenda += "</div>"

    return f"""
<section class="reveal animate">
  <h2 class="section-title">Semaforo por Entidad</h2>
  <p class="section-sub">Vista rapida (modo general) del desempeno de los modelos</p>
  <div class="card">
    {leyenda}
    <div class="table-wrapper">
      <table class="table-semaforo">
        <thead>{header}</thead>
        <tbody>{rows}</tbody>
      </table>
    </div>
  </div>
</section>"""


def _isnan(v: Any) -> bool:
    return isinstance(v, float) and math.isnan(v)


# ---------------------------------------------------------------------------
# Generacion HTML completa
# ---------------------------------------------------------------------------


def _generate_html(comp: pd.DataFrame, anio: int, semana: int) -> str:
    ahora = datetime.now(tz=TZ_MX).strftime("%d/%m/%Y %H:%M hrs")
    date_range = _epiweek_date_range(anio, semana)

    # KPIs
    n_series = len(comp)
    valid = comp[comp["error_pct"].notna()]
    n_valid = len(valid)
    n_good = len(valid[valid["error_pct"] <= 25])
    nac_gen = comp[(comp["nivel"] == "nacional") & (comp["sexo"] == "general")]
    nac_err = nac_gen["error_pct"].dropna()
    nac_mean = f"{nac_err.mean():.1f}%" if len(nac_err) > 0 else "N/A"

    # MASE medio nacional (CV del modelo ganador)
    prod_df = pd.read_excel(PROD_EXCEL, sheet_name=0)
    nac_mase = prod_df[(prod_df["entidad"] == "Nacional") & (prod_df["sexo"] == "general")][
        "mase_prod"
    ]
    mase_mean = f"{nac_mase.mean():.2f}" if len(nac_mase) > 0 else "N/A"

    css = _build_css()
    parts: list[str] = []
    parts.append(f"""<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Validacion Semanal - Semana {semana}/{anio}</title>
{_FONTS_LINK}
<style>{css}</style>
<link rel="stylesheet" href="editorial.css?v=1">
</head>
<body>
<div class="hero">
  <h1>Validacion Semanal: Realidad vs Pronostico</h1>
  <p class="subtitle">Semana Epidemiologica {semana} | {date_range}</p>
  <div class="hero-kpis">
    <div class="hero-kpi">
      <span class="value">{n_series}</span>
      <span class="label">Series evaluadas</span>
    </div>
    <div class="hero-kpi">
      <span class="value">{nac_mean}</span>
      <span class="label">SMAPE nacional medio</span>
    </div>
    <div class="hero-kpi">
      <span class="value">{mase_mean}</span>
      <span class="label">MASE nacional medio</span>
    </div>
    <div class="hero-kpi">
      <span class="value">{n_good}/{n_valid}</span>
      <span class="label">Precision &le;25%</span>
    </div>
  </div>
</div>
<div class="container">
""")

    parts.append(_section_dashboard(comp, anio, semana))
    parts.append(_section_nacional(comp))
    parts.append(_section_regional(comp))
    parts.append(_section_detalle(comp))
    parts.append(_section_semaforo(comp))

    parts.append(f"""
</div>
<footer>
  <p class="footer-title">EpiForecast-MX | IMSS 2026</p>
  <p>Datos: Boletin Epidemiologico SINAVE | Generado: {ahora}</p>
  <p>Modelos entrenados con datos hasta semana 1/2026</p>
</footer>
<script>
document.addEventListener('DOMContentLoaded',function(){{
  var els=document.querySelectorAll('.reveal');
  if(!els.length) return;
  var io=new IntersectionObserver(function(entries){{
    entries.forEach(function(e){{if(e.isIntersecting){{e.target.classList.add('visible');io.unobserve(e.target)}}}})
  }},{{threshold:0.1}});
  els.forEach(function(el){{io.observe(el)}});
}});
</script>
</body>
</html>""")

    return "".join(parts)


# ---------------------------------------------------------------------------
# CSS
# ---------------------------------------------------------------------------


def _build_css() -> str:
    return """
*,*::before,*::after{box-sizing:border-box;margin:0;padding:0}
:root{
  --burgundy:#F472B6;--dark-burgundy:#8E2A63;--teal:#5B8DEF;--dark-teal:#16203A;
  --gold:#2DD4BF;--cream:#E7ECF5;--cream-light:#1C2740;--cream-pale:#131C30;
  --cool-gray:#9FB0CE;--neutral-black:#0B0F1A;
  --burgundy-light:#F472B6;--teal-light:#5B8DEF;--gold-light:#2DD4BF;
  --orange:#FF6F00;--indigo:#5B8DEF;
  --font-display:'DM Serif Display',Georgia,serif;
  --font-body:'Source Sans 3','Source Sans Pro',sans-serif;
  --font-mono:'JetBrains Mono','Fira Code',monospace;
  --shadow-sm:0 2px 8px rgba(0,0,0,.06);--shadow-md:0 8px 24px rgba(0,0,0,.06);
  --shadow-lg:0 16px 48px rgba(0,0,0,.06);--radius:16px;--radius-sm:10px;
}
body{font-family:var(--font-body);background:var(--cream-pale);color:var(--neutral-black);
  line-height:1.7;-webkit-font-smoothing:antialiased}
body::before{content:'';position:fixed;inset:0;
  background-image:url("data:image/svg+xml,%3Csvg viewBox='0 0 256 256' xmlns='http://www.w3.org/2000/svg'%3E%3Cfilter id='n'%3E%3CfeTurbulence type='fractalNoise' baseFrequency='0.9' numOctaves='4' stitchTiles='stitch'/%3E%3C/filter%3E%3Crect width='100%25' height='100%25' filter='url(%23n)'/%3E%3C/svg%3E");
  opacity:.025;pointer-events:none;z-index:0}
.container{max-width:1200px;margin:0 auto;padding:0 2rem;position:relative;z-index:1}
section{padding:3rem 0}
.section-title{font-family:var(--font-display);font-size:clamp(1.8rem,4vw,2.5rem);
  margin-bottom:.5rem;color:var(--dark-teal)}
.section-sub{color:var(--cool-gray);font-size:1rem;margin-bottom:2rem}
.hero{padding:6rem 2rem 4rem;text-align:center;
  background:linear-gradient(170deg,var(--dark-teal) 0%,var(--teal) 60%,rgba(0,82,78,.85) 100%);
  color:var(--cream-pale);position:relative;overflow:hidden}
.hero::before{content:'';position:absolute;inset:0;
  background:radial-gradient(ellipse 800px 600px at 30% 20%,rgba(181,133,0,.12),transparent),
  radial-gradient(ellipse 600px 500px at 70% 80%,rgba(155,34,66,.08),transparent);pointer-events:none}
.hero h1{font-family:var(--font-display);font-size:clamp(2rem,5vw,3.2rem);line-height:1.15;
  margin-bottom:.75rem;position:relative}
.hero .subtitle{font-size:clamp(.95rem,1.8vw,1.15rem);opacity:.8;font-weight:300;
  margin-bottom:2.5rem;position:relative}
.hero-kpis{display:flex;justify-content:center;gap:1.5rem;flex-wrap:wrap;position:relative}
.hero-kpi{background:rgba(255,255,255,.08);border:1px solid rgba(232,213,181,.15);
  border-radius:var(--radius);padding:1.5rem 2rem;backdrop-filter:blur(8px);min-width:160px;
  transition:transform .3s,box-shadow .3s}
.hero-kpi:hover{transform:translateY(-4px);box-shadow:0 12px 32px rgba(0,0,0,.2)}
.hero-kpi .value{font-family:var(--font-display);font-size:2.5rem;display:block;
  background:linear-gradient(135deg,var(--cream),var(--gold-light));
  -webkit-background-clip:text;-webkit-text-fill-color:transparent;background-clip:text}
.hero-kpi .label{font-size:.8rem;text-transform:uppercase;letter-spacing:1px;opacity:.7}
.card{background:#fff;border-radius:var(--radius);padding:2rem;box-shadow:var(--shadow-md);
  border:1px solid rgba(0,0,0,.04);transition:transform .3s,box-shadow .3s;margin-bottom:2rem}
.card:hover{transform:translateY(-2px);box-shadow:var(--shadow-lg)}
.card h2{font-family:var(--font-display);font-size:1.5rem;margin-bottom:1rem;
  padding-bottom:.75rem;border-bottom:2px solid var(--cream-light);color:var(--dark-teal)}
.card-gold{border-left:4px solid var(--gold)}
.card-footer{display:flex;align-items:center;flex-wrap:wrap;gap:.75rem;
  margin-top:1rem;padding-top:1rem;border-top:1px solid var(--cream-light)}
.table-wrapper{overflow-x:auto;margin-top:.5rem}
table{width:100%;border-collapse:collapse;font-size:.88rem}
thead th{padding:.85rem 1rem;text-align:center;font-weight:600;text-transform:uppercase;
  font-size:.72rem;letter-spacing:.5px;color:var(--cool-gray);
  border-bottom:2px solid var(--cream-light);white-space:nowrap;
  position:sticky;top:0;background:#fff;z-index:2}
td:first-child,th:first-child{text-align:left}
tbody td{padding:.7rem 1rem;border-bottom:1px solid var(--cream-light);vertical-align:middle;
  font-family:var(--font-mono);font-size:.82rem;text-align:center}
tbody td:first-child{font-family:var(--font-body);font-size:.88rem;text-align:left}
tbody tr{transition:background .15s}
tbody tr:hover{background:var(--cream-pale)}
.row-nacional{background:rgba(181,133,0,.06);font-weight:700}
.row-nacional td{border-bottom:2px solid var(--gold)}
.row-group-last td{border-bottom:2px solid var(--cream)}
.err-good{color:#2DD4BF;font-weight:600}
.err-bad{color:#EF4444;font-weight:600}
.mase-good{color:#2DD4BF;font-weight:600}
.mase-warn{color:#EF4444;font-weight:600}
.table-semaforo td{text-align:center;padding:.5rem .75rem}
.table-semaforo th{min-width:140px}
.leyenda{display:flex;gap:1.5rem;flex-wrap:wrap;margin-bottom:1.5rem;
  padding:1rem;background:var(--cream-pale);border-radius:var(--radius-sm)}
.leyenda-item{display:inline-flex;align-items:center;gap:.3rem;font-size:.85rem}
.reveal.animate{opacity:0;transform:translateY(30px);transition:opacity .6s ease,transform .6s ease}
.reveal.visible{opacity:1;transform:translateY(0)}

.dashboard{background:linear-gradient(170deg,#0a1628 0%,var(--dark-teal) 50%,#0d2b28 100%);
  border-radius:var(--radius);padding:3rem 2rem;margin-bottom:2rem;position:relative;overflow:hidden}
.dashboard::before{content:'';position:absolute;inset:0;
  background:radial-gradient(ellipse 600px 400px at 20% 30%,rgba(181,133,0,.08),transparent),
  radial-gradient(ellipse 400px 300px at 80% 70%,rgba(155,34,66,.06),transparent);pointer-events:none}
.dashboard::after{content:'';position:absolute;inset:0;
  background-image:url("data:image/svg+xml,%3Csvg viewBox='0 0 256 256' xmlns='http://www.w3.org/2000/svg'%3E%3Cfilter id='n'%3E%3CfeTurbulence type='fractalNoise' baseFrequency='.8' numOctaves='4' stitchTiles='stitch'/%3E%3C/filter%3E%3Crect width='100%25' height='100%25' filter='url(%23n)'/%3E%3C/svg%3E");
  opacity:.03;pointer-events:none}
.dashboard-header{text-align:center;margin-bottom:2.5rem;position:relative;z-index:1}
.dashboard-header h2{font-family:var(--font-display);font-size:clamp(1.6rem,3.5vw,2.2rem);
  color:var(--cream);margin-bottom:.4rem}
.dashboard-header p{color:rgba(232,213,181,.5);font-size:.95rem}
.dashboard-grid{display:grid;grid-template-columns:repeat(3,1fr);gap:1.5rem;position:relative;z-index:1}
.dash-panel{background:rgba(255,255,255,.05);border:1px solid rgba(232,213,181,.1);
  border-radius:var(--radius);padding:2rem 1.5rem;text-align:center;
  backdrop-filter:blur(12px);transition:transform .3s,box-shadow .3s}
.dash-panel:hover{transform:translateY(-5px);box-shadow:0 20px 50px rgba(0,0,0,.35)}
.dash-panel h3{font-family:var(--font-display);color:var(--cream);font-size:1.1rem;
  margin-bottom:1.2rem;letter-spacing:.3px}
.gauge-wrap{display:flex;justify-content:center;margin-bottom:1.2rem}
.gauge-arc{stroke-dashoffset:408.41;transition:stroke-dashoffset 2s cubic-bezier(.4,0,.2,1) .6s}
.visible .gauge-arc{stroke-dashoffset:var(--to)}
.dash-nums{display:flex;align-items:center;justify-content:center;gap:.8rem;margin-bottom:1.2rem}
.dash-n{text-align:center}
.dash-v{font-family:var(--font-display);font-size:1.4rem;color:var(--cream);display:block}
.dash-l{font-size:.7rem;color:rgba(232,213,181,.45);text-transform:uppercase;letter-spacing:.5px}
.dash-badges{display:flex;justify-content:center;gap:.5rem;margin-top:1rem}
.mase-box{margin:0 0 .5rem;padding:1rem;background:rgba(0,0,0,.18);border-radius:var(--radius-sm)}
.mase-head{display:flex;justify-content:space-between;align-items:center;margin-bottom:.5rem}
.mase-title{color:rgba(232,213,181,.5);font-size:.75rem;text-transform:uppercase;letter-spacing:.5px}
.mase-val{font-family:var(--font-display);font-size:1.15rem}
.mase-track{height:6px;background:rgba(255,255,255,.07);border-radius:3px;position:relative;
  margin-bottom:.3rem}
.mase-fill{height:100%;border-radius:3px;width:0;
  transition:width 1.8s cubic-bezier(.4,0,.2,1) 1.2s}
.visible .mase-fill{width:var(--mw)}
.mase-ref{position:absolute;left:50%;top:-4px;width:2px;height:14px;
  background:rgba(232,213,181,.35);border-radius:1px}
.mase-ticks{display:flex;justify-content:space-between;position:relative;
  font-size:.6rem;color:rgba(232,213,181,.25);font-family:var(--font-mono)}
.mase-lbl{font-size:.78rem;font-weight:600;margin-top:.3rem}

footer{background:var(--dark-teal);color:rgba(232,213,181,.7);padding:3rem 2rem;
  text-align:center;font-size:.85rem;position:relative;z-index:1;margin-top:2rem}
footer .footer-title{font-family:var(--font-display);color:var(--cream);font-size:1.2rem;
  margin-bottom:.5rem}
@media(max-width:768px){
  .hero{padding:4rem 1.5rem 3rem}.hero-kpis{gap:.75rem}
  .hero-kpi{min-width:130px;padding:1rem 1.25rem}.hero-kpi .value{font-size:1.8rem}
  .container{padding:0 1rem}section{padding:2rem 0}
  .card-footer{flex-direction:column;align-items:flex-start}
  .leyenda{flex-direction:column;gap:.5rem}
  .dashboard-grid{grid-template-columns:1fr}.dashboard{padding:2rem 1rem}
}
"""


# ---------------------------------------------------------------------------
# Actualizacion del Excel de produccion
# ---------------------------------------------------------------------------


def _update_excel(comp: pd.DataFrame, anio: int, semana: int) -> None:
    """Actualiza realidad_sem_previa en el Excel de produccion."""
    from openpyxl import load_workbook

    wb = load_workbook(PROD_EXCEL)
    ws = wb.active
    if ws is None:
        print("  Error: no se pudo abrir la hoja del Excel")
        return

    # Indices de columnas (1-indexed)
    headers = [cell.value for cell in ws[1]]
    try:
        col_real = headers.index("realidad_sem_previa") + 1
        col_pad = headers.index("padecimiento") + 1
        col_ent = headers.index("entidad") + 1
        col_sex = headers.index("sexo") + 1
    except ValueError as e:
        print(f"  Error: columna no encontrada en Excel — {e}")
        return

    # Lookup de realidad
    lookup: dict[tuple[str, str, str], int | None] = {}
    for _, row in comp.iterrows():
        raw = row["real"]
        val = int(round(raw)) if raw is not None and not _isnan(raw) else None
        lookup[(row["pad_norm"], row["entidad_norm"], row["sexo"])] = val

    updated = 0
    for row_num in range(2, ws.max_row + 1):
        pad_raw = str(ws.cell(row=row_num, column=col_pad).value or "")
        ent_raw = str(ws.cell(row=row_num, column=col_ent).value or "")
        sex_raw = str(ws.cell(row=row_num, column=col_sex).value or "")

        pad = _normalize(pad_raw)
        ent = _norm_prod_entidad(ent_raw)

        key = (pad, ent, sex_raw)
        if key in lookup and lookup[key] is not None:
            ws.cell(row=row_num, column=col_real).value = lookup[key]
            updated += 1

    wb.save(PROD_EXCEL)
    print(f"  Excel actualizado: {updated}/333 filas con realidad S{semana}/{anio}")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------


def main() -> None:
    print("Cargando boletin epidemiologico...")
    boletin, anio, semana = _load_boletin_full()
    n_states = boletin["entidad_norm"].nunique()
    print(f"  Semana {anio}-S{semana} | {n_states} entidades x 3 sexos = {len(boletin)} filas")

    target_date = _epiweek_to_monday(anio, semana).isoformat()
    print(f"  Fecha forecast: {target_date}")

    print("Cargando forecasts (4 modelos, todos los modos)...")
    forecasts = _load_forecasts(target_date)
    print(f"  {len(forecasts)} registros")

    print("Cargando tabla de produccion...")
    produccion = _load_production()
    print(f"  {len(produccion)} modelos")

    print("Construyendo comparacion...")
    comp = _build_comparison(boletin, forecasts, produccion)
    print(f"  {len(comp)} series ({comp['nivel'].value_counts().to_dict()})")

    # Estadisticas
    for nivel in ["nacional", "regional", "estatal"]:
        niv = comp[comp["nivel"] == nivel]
        valid = niv[niv["error_pct"].notna()]
        n_good = len(valid[valid["error_pct"] <= 25])
        print(f"  {nivel}: {n_good}/{len(valid)} con error <=25%")

    print("Generando HTML...")
    html = _generate_html(comp, anio, semana)
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT.write_text(html, encoding="utf-8")
    print(f"  {OUTPUT} ({len(html):,} bytes)")

    print("Actualizando Excel de produccion...")
    _update_excel(comp, anio, semana)

    print("Listo.")


if __name__ == "__main__":
    main()
