"""Formato Excel IMSS para la tabla de produccion.

Aplica estilos institucionales (paleta IMSS 2026) a las hojas del workbook.
"""

from __future__ import annotations

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

# ---------------------------------------------------------------------------
# Paleta IMSS 2026
# ---------------------------------------------------------------------------
_VERDE_IMSS = "006847"
_VERDE_CLARO = "E8F5E9"
_GRIS_HEADER = "F5F5F5"
_BLANCO = "FFFFFF"
_NEGRO = "212121"
_ROJO = "C62828"
_AMARILLO = "FFF8E1"
_NARANJA = "FF6F00"
_AZUL_SUAVE = "E3F2FD"
_BORDE_COLOR = "BDBDBD"

_HEADER_FILL = PatternFill("solid", fgColor=_VERDE_IMSS)
_HEADER_FONT = Font(name="Calibri", size=10, bold=True, color=_BLANCO)
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
_DATA_FONT = Font(name="Calibri", size=9, color=_NEGRO)
_DATA_ALIGN = Alignment(horizontal="center", vertical="center")
_DATA_ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
_THIN_BORDER = Border(
    left=Side(style="thin", color=_BORDE_COLOR),
    right=Side(style="thin", color=_BORDE_COLOR),
    top=Side(style="thin", color=_BORDE_COLOR),
    bottom=Side(style="thin", color=_BORDE_COLOR),
)
_FILL_EVEN = PatternFill("solid", fgColor=_VERDE_CLARO)
_FILL_ODD = PatternFill("solid", fgColor=_BLANCO)

# Columnas que se alinean a la izquierda
_LEFT_COLS = {
    "padecimiento",
    "entidad",
    "sexo",
    "justificacion",
    "overfitting",
    "leakage",
    "modelo_produccion",
    "tipo_modelo",
    "region_asignada",
}


def format_sheet(ws: Worksheet, freeze_col: int = 4) -> None:
    """Aplica formato IMSS a una hoja ya llena de datos."""
    max_row = ws.max_row
    max_col = ws.max_column

    # 1. Headers (fila 1)
    for col_idx in range(1, max_col + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = _HEADER_ALIGN
        cell.border = _THIN_BORDER

    # Leer nombres de columna de la fila 1
    col_names: dict[int, str] = {}
    for col_idx in range(1, max_col + 1):
        val = ws.cell(row=1, column=col_idx).value
        col_names[col_idx] = str(val).lower() if val else ""

    # 2. Data rows (filas 2+)
    for row_idx in range(2, max_row + 1):
        fill = _FILL_EVEN if row_idx % 2 == 0 else _FILL_ODD
        for col_idx in range(1, max_col + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = _DATA_FONT
            cell.border = _THIN_BORDER
            cell.fill = fill
            name = col_names.get(col_idx, "")
            if name in _LEFT_COLS:
                cell.alignment = _DATA_ALIGN_LEFT
            else:
                cell.alignment = _DATA_ALIGN

    # 3. Anchos de columna automaticos
    for col_idx in range(1, max_col + 1):
        name = col_names.get(col_idx, "")
        # Base: longitud del header + 2
        header_len = len(str(ws.cell(row=1, column=col_idx).value or ""))
        width = max(header_len + 2, 10)
        # Ajustar columnas especificas
        if name == "justificacion":
            width = 55
        elif name in ("padecimiento", "entidad"):
            width = 22
        elif name in ("modelo_produccion", "region_asignada"):
            width = 18
        elif name in ("overfitting", "leakage"):
            width = 20
        elif "semana" in name or "sem_" in name or name.startswith(("real_", "pron_", "acierto_")):
            width = 8
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # 4. Congelar paneles (header + primeras N columnas)
    ws.freeze_panes = ws.cell(row=2, column=freeze_col + 1)

    # 5. Filtro automatico
    ws.auto_filter.ref = f"A1:{get_column_letter(max_col)}{max_row}"


def format_detail_header(ws: Worksheet) -> None:
    """Aplica colores diferenciados a los bloques de la hoja de detalle semanal.

    - Columnas base: verde IMSS (igual que format_sheet)
    - Bloque real_sem_*: azul suave
    - Bloque pron_sem_*: amarillo suave
    - Bloque acierto_sem_*: gris header
    """
    _fill_blue = PatternFill("solid", fgColor="1565C0")
    _fill_orange = PatternFill("solid", fgColor=_NARANJA)
    _fill_gray = PatternFill("solid", fgColor="616161")

    max_col = ws.max_column
    for col_idx in range(1, max_col + 1):
        cell = ws.cell(row=1, column=col_idx)
        name = str(cell.value or "").lower()
        if name.startswith("real_"):
            cell.fill = _fill_blue
        elif name.startswith("pron_"):
            cell.fill = _fill_orange
        elif name.startswith("acierto_"):
            cell.fill = _fill_gray
        # Los demas ya tienen _HEADER_FILL del format_sheet
