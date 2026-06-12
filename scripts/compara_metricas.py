# scripts/compara_metricas.py
"""Genera un Excel comparativo de metricas entre todos los modelos disponibles.

Soporta: Prophet, DeepAR, Ensemble, Stacking (solo incluye los que tengan datos).

Uso:
    python -m scripts.compara_metricas
    make compare-metrics
"""

from __future__ import annotations

from datetime import datetime
from pathlib import Path
from zoneinfo import ZoneInfo

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
import pandas as pd

from epiforecast.utils.config import logger
from epiforecast.visualization.comparison_report import generar_reporte_html

# ---------------------------------------------------------------------------
# Constantes
# ---------------------------------------------------------------------------
MODELS: dict[str, dict[str, str]] = {
    "prophet": {"color": "004D40", "label": "Prophet"},
    "deepar": {"color": "880E4F", "label": "DeepAR"},
    "ensemble": {"color": "FF6F00", "label": "Ensemble"},
    "stacking": {"color": "1A237E", "label": "Stacking"},
}

METRIC_COLS = ["rmse", "mae", "mape", "smape", "mase"]
INFO_COLS = ["confianza", "tiempo_total_seg"]
MERGE_KEYS = ["padecimiento", "sexo", "nivel", "Entidad"]

DARK_GRAY = "424242"
GREEN_BG = "C8E6C9"
RED_BG = "FFCDD2"
WHITE_FG = "FFFFFF"
GANADOR_BG = "37474F"

MODELS_DIR = Path("models")
OUTPUT_DIR = Path("reports/forecasts/comparacion_modelos")
OUTPUT_FILE = OUTPUT_DIR / "comparacion_metricas.xlsx"
TZ_CDMX = ZoneInfo("America/Mexico_City")

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


# ---------------------------------------------------------------------------
# Lectura de datos
# ---------------------------------------------------------------------------
def _leer_completos(modelo: str) -> pd.DataFrame:
    """Lee todos los CSV ``*_completo.csv`` de un modelo y los concatena."""
    base = MODELS_DIR / modelo
    archivos = sorted(base.rglob("*_completo.csv"))
    if not archivos:
        return pd.DataFrame()
    frames: list[pd.DataFrame] = []
    for f in archivos:
        logger.info("  Leyendo {}", f.relative_to("."))
        frames.append(pd.read_csv(f))
    return pd.concat(frames, ignore_index=True)


def _normalizar_keys(df: pd.DataFrame) -> pd.DataFrame:
    """Agrega columna ``Entidad`` si no existe y rellena NaN."""
    if "Entidad" not in df.columns:
        df["Entidad"] = ""
    df["Entidad"] = df["Entidad"].fillna("")
    if "nivel" not in df.columns:
        df["nivel"] = "nacional"
    df["nivel"] = df["nivel"].fillna("nacional")
    return df


def _cargar_modelos() -> dict[str, pd.DataFrame]:
    """Carga datos de todos los modelos disponibles."""
    data: dict[str, pd.DataFrame] = {}
    for nombre in MODELS:
        logger.info("Leyendo metricas {}...", MODELS[nombre]["label"])
        df = _leer_completos(nombre)
        if not df.empty:
            data[nombre] = df
            logger.info("  {} filas cargadas", len(df))
        else:
            logger.warning("  Sin datos para {}", MODELS[nombre]["label"])
    return data


# ---------------------------------------------------------------------------
# Construccion de hojas
# ---------------------------------------------------------------------------
def _build_detalle(data: dict[str, pd.DataFrame]) -> pd.DataFrame:
    """Merge N-way: una columna por metrica por modelo + ganador."""
    # Recopilar todas las claves unicas
    all_keys = pd.DataFrame()
    for df in data.values():
        keys_df = _normalizar_keys(df.copy())[MERGE_KEYS].drop_duplicates()
        all_keys = pd.concat([all_keys, keys_df]).drop_duplicates()
    result = all_keys.reset_index(drop=True)

    # Merge cada modelo
    for nombre, df in data.items():
        df_norm = _normalizar_keys(df.copy())
        cols_to_keep = MERGE_KEYS.copy()
        available_metrics = [c for c in METRIC_COLS if c in df_norm.columns]
        available_info = [c for c in INFO_COLS if c in df_norm.columns]
        cols_to_keep += available_metrics + available_info
        df_subset = df_norm[[c for c in cols_to_keep if c in df_norm.columns]].copy()
        rename_map = {c: f"{c}_{nombre}" for c in available_metrics + available_info}
        df_subset = df_subset.rename(columns=rename_map)
        result = result.merge(df_subset, on=MERGE_KEYS, how="left")

    # Ganador por metrica (valor minimo = mejor)
    model_names = list(data.keys())
    for m in METRIC_COLS:
        model_cols = {n: f"{m}_{n}" for n in model_names if f"{m}_{n}" in result.columns}
        if len(model_cols) < 2:
            continue
        ganador_vals = []
        for _, row in result.iterrows():
            vals = {n: row.get(col) for n, col in model_cols.items()}
            valid = {n: v for n, v in vals.items() if pd.notna(v)}
            if not valid:
                ganador_vals.append("N/D")
            else:
                best = min(valid, key=lambda k: valid[k])
                ganador_vals.append(MODELS[best]["label"])
        result[f"ganador_{m}"] = ganador_vals

    # Ordenar columnas
    ordered: list[str] = list(MERGE_KEYS)
    for m in METRIC_COLS:
        for n in model_names:
            col = f"{m}_{n}"
            if col in result.columns:
                ordered.append(col)
        g = f"ganador_{m}"
        if g in result.columns:
            ordered.append(g)
    for n in model_names:
        for info in INFO_COLS:
            col = f"{info}_{n}"
            if col in result.columns:
                ordered.append(col)
    ordered += [c for c in result.columns if c not in ordered]
    return result[ordered].sort_values(MERGE_KEYS, ignore_index=True)


def _build_resumen(detalle: pd.DataFrame, model_names: list[str]) -> pd.DataFrame:
    """Tabla resumen: promedio de metricas por padecimiento y modelo."""
    rows: list[dict[str, object]] = []
    for pad, grp in detalle.groupby("padecimiento"):
        row: dict[str, object] = {"padecimiento": pad}
        for m in METRIC_COLS:
            best_val = float("inf")
            best_name = "N/D"
            for n in model_names:
                col = f"{m}_{n}"
                if col in grp.columns:
                    val = grp[col].mean(skipna=True)
                    row[col] = val
                    if pd.notna(val) and val < best_val:
                        best_val = val
                        best_name = MODELS[n]["label"]
            row[f"ganador_{m}"] = best_name
        rows.append(row)

    ordered_cols = ["padecimiento"]
    for m in METRIC_COLS:
        for n in model_names:
            ordered_cols.append(f"{m}_{n}")
        ordered_cols.append(f"ganador_{m}")
    df = pd.DataFrame(rows)
    return df[[c for c in ordered_cols if c in df.columns]]


def _build_metadata(
    all_files: dict[str, list[Path]],
) -> pd.DataFrame:
    """Hoja de metadata con informacion de generacion."""
    ahora = datetime.now(tz=TZ_CDMX)
    campos = [
        ("Fecha de generacion", ahora.strftime("%Y-%m-%d %H:%M:%S")),
        ("Zona horaria", "America/Mexico_City (UTC-6)"),
        ("Version del proyecto", "2.0.0"),
    ]
    for nombre, files in all_files.items():
        label = MODELS[nombre]["label"]
        campos.append((f"Archivos {label}", ", ".join(str(f.relative_to(".")) for f in files)))
    campos.append(("Directorio de salida", str(OUTPUT_FILE)))
    return pd.DataFrame(campos, columns=["Campo", "Valor"])


# ---------------------------------------------------------------------------
# Nombres de columnas legibles
# ---------------------------------------------------------------------------
def _build_col_display(model_names: list[str]) -> dict[str, str]:
    """Genera mapeo de nombres internos a nombres legibles."""
    display: dict[str, str] = {
        "padecimiento": "Padecimiento",
        "sexo": "Sexo",
        "nivel": "Nivel",
        "Entidad": "Entidad",
    }
    for n in model_names:
        label = MODELS[n]["label"]
        for m in METRIC_COLS:
            display[f"{m}_{n}"] = f"{m.upper()} {label}"
        display[f"confianza_{n}"] = f"Confianza {label}"
        display[f"tiempo_total_seg_{n}"] = f"Tiempo (s) {label}"
    for m in METRIC_COLS:
        display[f"ganador_{m}"] = f"Mejor {m.upper()}"
    return display


# ---------------------------------------------------------------------------
# Estilos openpyxl
# ---------------------------------------------------------------------------
def _autofit_columns(ws) -> None:  # type: ignore[no-untyped-def]
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 35)


def _aplicar_bordes(ws, max_row: int, max_col: int) -> None:  # type: ignore[no-untyped-def]
    for row in range(1, max_row + 1):
        for col in range(1, max_col + 1):
            ws.cell(row=row, column=col).border = THIN_BORDER


def _aplicar_formato_numerico(ws, df: pd.DataFrame, start_row: int = 2) -> None:  # type: ignore[no-untyped-def]
    for col_idx, col_name in enumerate(df.columns, 1):
        col_lower = col_name.lower()
        if any(k in col_lower for k in ("rmse", "mae", "mase")) and "mejor" not in col_lower:
            fmt = "0.0000"
        elif "mape" in col_lower or "smape" in col_lower:
            fmt = "0.00"
        elif "tiempo" in col_lower:
            fmt = "0.1"
        else:
            continue
        for row in range(start_row, start_row + len(df)):
            ws.cell(row=row, column=col_idx).number_format = fmt


def _aplicar_condicional_detalle(ws, df: pd.DataFrame, start_row: int = 2) -> None:  # type: ignore[no-untyped-def]
    """Verde si MASE < 1, rojo si confianza = 'insuficiente'."""
    green_fill = PatternFill(start_color=GREEN_BG, end_color=GREEN_BG, fill_type="solid")
    red_fill = PatternFill(start_color=RED_BG, end_color=RED_BG, fill_type="solid")
    for col_idx, col_name in enumerate(df.columns, 1):
        col_lower = col_name.lower()
        is_mase = "mase" in col_lower and "mejor" not in col_lower
        is_confianza = "confianza" in col_lower
        if not is_mase and not is_confianza:
            continue
        for row_idx, val in enumerate(df[col_name], start_row):
            cell = ws.cell(row=row_idx, column=col_idx)
            if is_mase and pd.notna(val) and isinstance(val, int | float) and val < 1:
                cell.fill = green_fill
            if is_confianza and val == "insuficiente":
                cell.fill = red_fill


def _header_color(col_name: str) -> str:
    col_lower = col_name.lower()
    for cfg in MODELS.values():
        if cfg["label"].lower() in col_lower:
            return cfg["color"]
    if "mejor" in col_lower:
        return GANADOR_BG
    return DARK_GRAY


def _estilizar_headers(ws, df: pd.DataFrame) -> None:  # type: ignore[no-untyped-def]
    for col_idx, col_name in enumerate(df.columns, 1):
        color = _header_color(col_name)
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        cell.font = Font(bold=True, color=WHITE_FG, size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER


def _colorear_ganadores(ws, df: pd.DataFrame) -> None:  # type: ignore[no-untyped-def]
    for col_idx, col_name in enumerate(df.columns, 1):
        if "mejor" not in col_name.lower():
            continue
        for row_idx, val in enumerate(df[col_name], 2):
            cell = ws.cell(row=row_idx, column=col_idx)
            for cfg in MODELS.values():
                if val == cfg["label"]:
                    cell.font = Font(bold=True, color=cfg["color"])
                    break


def _estilo_metadata(ws) -> None:  # type: ignore[no-untyped-def]
    fill = PatternFill(start_color=DARK_GRAY, end_color=DARK_GRAY, fill_type="solid")
    font = Font(bold=True, color=WHITE_FG, size=11)
    for col in range(1, 3):
        cell = ws.cell(row=1, column=col)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER


# ---------------------------------------------------------------------------
# Generacion del Excel
# ---------------------------------------------------------------------------
def generar_excel_comparativo() -> Path:
    """Genera el Excel comparativo y devuelve la ruta del archivo."""
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    data = _cargar_modelos()
    if not data:
        logger.error("No se encontraron datos de ningun modelo.")
        raise SystemExit(1)

    model_names = list(data.keys())
    labels = [MODELS[n]["label"] for n in model_names]
    logger.info("Modelos encontrados: {}", ", ".join(labels))

    # Archivos por modelo (para metadata)
    all_files: dict[str, list[Path]] = {}
    for n in model_names:
        all_files[n] = sorted((MODELS_DIR / n).rglob("*_completo.csv"))

    # Construir hojas
    logger.info("Construyendo hoja Detalle...")
    detalle = _build_detalle(data)

    logger.info("Construyendo hoja Resumen...")
    resumen = _build_resumen(detalle, model_names)

    logger.info("Construyendo hoja Metadata...")
    metadata = _build_metadata(all_files)

    # Renombrar columnas
    col_display = _build_col_display(model_names)
    resumen_xl = resumen.rename(columns=col_display)
    detalle_xl = detalle.rename(columns=col_display)

    # Escribir Excel
    logger.info("Escribiendo Excel en {}...", OUTPUT_FILE)
    with pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl") as writer:
        resumen_xl.to_excel(writer, sheet_name="Resumen", index=False)
        detalle_xl.to_excel(writer, sheet_name="Detalle", index=False)
        metadata.to_excel(writer, sheet_name="Metadata", index=False)

        wb = writer.book

        # Estilizar Resumen
        ws_res = wb["Resumen"]
        _estilizar_headers(ws_res, resumen_xl)
        _aplicar_formato_numerico(ws_res, resumen_xl)
        _colorear_ganadores(ws_res, resumen_xl)
        _aplicar_bordes(ws_res, len(resumen_xl) + 1, len(resumen_xl.columns))
        _autofit_columns(ws_res)
        ws_res.freeze_panes = "B2"

        # Estilizar Detalle
        ws_det = wb["Detalle"]
        _estilizar_headers(ws_det, detalle_xl)
        _aplicar_formato_numerico(ws_det, detalle_xl)
        _aplicar_condicional_detalle(ws_det, detalle_xl, start_row=2)
        _aplicar_bordes(ws_det, len(detalle_xl) + 1, len(detalle_xl.columns))
        _autofit_columns(ws_det)
        ws_det.freeze_panes = "E2"

        # Estilizar Metadata
        ws_meta = wb["Metadata"]
        _estilo_metadata(ws_meta)
        _aplicar_bordes(ws_meta, len(metadata) + 1, 2)
        _autofit_columns(ws_meta)

    logger.success("Excel generado: {}", OUTPUT_FILE)
    return OUTPUT_FILE


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------
def main() -> None:
    logger.info("Iniciando comparacion de metricas multi-modelo...")
    generar_excel_comparativo()
    generar_reporte_html()
    logger.success("Proceso de comparacion de metricas finalizado.")


if __name__ == "__main__":
    main()
