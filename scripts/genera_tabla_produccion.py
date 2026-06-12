"""Genera Excel con la tabla completa de modelos y seleccion de produccion.

Compara los 4 algoritmos (Prophet, DeepAR, Ensemble, Stacking) para cada
combinacion (padecimiento x entidad x sexo) y selecciona el modelo de
produccion con justificacion automatica.

Hojas:
  - Produccion: 333 filas, metricas, diagnosticos, comparativa historica.
  - Detalle Semanal: 52 semanas de realidad vs pronostico + % acierto.

Uso:
    python -m scripts.genera_tabla_produccion
"""

from __future__ import annotations

from pathlib import Path

import numpy as np
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd

from epiforecast.utils.config import logger
from epiforecast.visualization.avance5_tables import (
    _MODEL_LABELS,
    cargar_completos,
    merge_all_models,
)

_MODELS = ["prophet", "deepar", "ensemble", "stacking"]
_METRICS = ["rmse", "mae", "smape", "mase"]
_EMPATE_PCT = 0.05  # 5% para considerar empate en SMAPE
_HORIZON = 52  # semanas de pronostico
_LOW_VOLUME_THRESHOLD = 5  # casos/52sem para considerar baja confianza

_SEXO_DISPLAY = {
    "incrementos_total": "general",
    "incrementos_hombres": "hombres",
    "incrementos_mujeres": "mujeres",
}

_MODO_TO_INCREMENTO = {
    "general": "incrementos_total",
    "hombres": "incrementos_hombres",
    "mujeres": "incrementos_mujeres",
}

# Umbrales diagnosticos (mismos que comparison_html.py)
_OVERFIT_ALTO = 2.0
_OVERFIT_MODERADO = 1.3
_LEAKAGE_THRESHOLD = 0.5

_OUTPUT = Path("reports") / "ProdDetails" / "tabla_333_modelos_produccion.xlsx"

_MODEL_KEY_MAP: dict[str, str] = {
    "Prophet": "prophet",
    "DeepAR": "deepar",
    "Ensemble": "ensemble",
    "Stacking": "stacking",
}

_SEXO_TO_MODO: dict[str, str] = {
    "general": "general",
    "hombres": "hombres",
    "mujeres": "mujeres",
}

_ZERO_THRESHOLD = 1e-6

# Cache de backtest DeepAR (se llena una vez con _build_deepar_backtest_cache)
_deepar_backtest: dict[tuple[str, str, str], pd.Series] = {}  # type: ignore[type-arg]


# ---------------------------------------------------------------------------
# Backtest DeepAR (resuelve fitted values reales, no copias del historico)
# ---------------------------------------------------------------------------


def _build_deepar_backtest_cache() -> dict[tuple[str, str, str], pd.Series]:  # type: ignore[type-arg]
    """Carga cada pkl DeepAR, recorta 52 sem, predice y retorna cache.

    Returns dict[(padecimiento, entidad, modo)] -> pd.Series[ds -> yhat].
    """
    from epiforecast.models.factory import create_model
    from epiforecast.utils.config import conf

    cache: dict[tuple[str, str, str], pd.Series] = {}  # type: ignore[type-arg]
    base = Path("models") / "deepar"
    if not base.exists():
        return cache

    pkls = sorted(base.rglob("*.pkl"))
    logger.info("DeepAR backtest: cargando {} modelos...", len(pkls))
    ok = 0
    fail = 0
    for pkl_path in pkls:
        try:
            model = create_model("deepar", config=conf)
            model.load(pkl_path)
            serie_full = model.serie.copy()
            if len(serie_full) < _HORIZON * 2:
                continue
            # Recortar ultimas 52 semanas
            model.serie = serie_full.iloc[:-_HORIZON].copy()
            forecast_df = model.predict(horizon=_HORIZON)
            # Las ultimas 52 filas del forecast son la prediccion real
            pred_52 = forecast_df.tail(_HORIZON).copy()
            pred_52["ds"] = pd.to_datetime(pred_52["ds"])

            # Extraer metadata del nombre: Deepar_{Pad}_{Entidad}_{modo}.pkl
            parts = pkl_path.stem.split("_")
            # parts[0] = "Deepar", parts[1] = padecimiento, parts[-1] = modo
            modo = parts[-1]  # general / hombres / mujeres
            pad = parts[1]  # Depresion / Alzheimer / Parkinson
            entidad_parts = parts[2:-1]
            entidad = " ".join(entidad_parts) if entidad_parts else "Nacional"

            cache[(pad, entidad, modo)] = pred_52.set_index("ds")["yhat"]
            ok += 1
        except Exception as exc:  # noqa: BLE001
            fail += 1
            logger.debug("  Backtest fallo {}: {}", pkl_path.name, exc)

    logger.info("DeepAR backtest completado: {}/{} exitosos, {} fallos", ok, ok + fail, fail)
    return cache


def _strip_accents(s: str) -> str:
    """Quita acentos comunes del español."""
    return (
        s.replace("ó", "o")
        .replace("á", "a")
        .replace("é", "e")
        .replace("í", "i")
        .replace("ú", "u")
        .replace("ñ", "n")
    )


def _get_deepar_backtest(padecimiento: str, entidad: str, sexo: str) -> pd.Series:  # type: ignore[type-arg]
    """Busca la serie de backtest DeepAR en el cache global."""
    modo = _SEXO_TO_MODO.get(sexo, sexo)
    pad_key = _strip_accents(padecimiento)
    # Normalizar entidad: quitar acentos y adaptar regiones
    ent_key = _strip_accents(entidad)
    # Regiones: "region_Urbana media" -> "region Urbana media" (pkl usa espacio)
    # "region_Rural / dispersa" -> "region Rural - dispersa" (pkl usa "-")
    if ent_key.startswith("region_"):
        ent_key = "region " + ent_key[len("region_") :].replace(" / ", " - ")

    key = (pad_key, ent_key, modo)
    if key in _deepar_backtest:
        return _deepar_backtest[key]
    # Busqueda fuzzy por si hay diferencias menores
    for (p, e, m), serie in _deepar_backtest.items():
        if p.lower() == pad_key.lower() and e.lower() == ent_key.lower() and m == modo:
            return serie
    return pd.Series(dtype=float)


# ---------------------------------------------------------------------------
# Funciones auxiliares de seleccion
# ---------------------------------------------------------------------------


def _best_model_for_metric(
    row: pd.Series,  # type: ignore[type-arg]
    metric: str,
    model_keys: list[str],
) -> str:
    vals: dict[str, float] = {}
    for mk in model_keys:
        v = row.get(f"{metric}_{mk}")
        if pd.notna(v):
            vals[mk] = float(v)
    if not vals:
        return ""
    best_key = min(vals, key=lambda k: vals[k])
    return _MODEL_LABELS.get(best_key, best_key)


def _count_wins(row: pd.Series, model_keys: list[str]) -> dict[str, int]:  # type: ignore[type-arg]
    wins: dict[str, int] = {mk: 0 for mk in model_keys}
    for metric in _METRICS:
        vals: dict[str, float] = {}
        for mk in model_keys:
            v = row.get(f"{metric}_{mk}")
            if pd.notna(v):
                vals[mk] = float(v)
        if vals:
            best = min(vals, key=lambda k: vals[k])
            wins[best] += 1
    return wins


def _get_metric_vals(
    row: pd.Series,  # type: ignore[type-arg]
    metric: str,
    available: list[str],
) -> dict[str, float]:
    vals: dict[str, float] = {}
    for mk in available:
        v = row.get(f"{metric}_{mk}")
        if pd.notna(v):
            vals[mk] = float(v)
    return vals


def _select_production(
    row: pd.Series,  # type: ignore[type-arg]
    model_keys: list[str],
) -> tuple[str, int, str]:
    """SMAPE primario, MASE desempate, RMSE segundo desempate."""
    available = [mk for mk in model_keys if pd.notna(row.get(f"smape_{mk}"))]
    if not available:
        return ("", 0, "Sin datos disponibles.")

    smape_vals = _get_metric_vals(row, "smape", available)
    mase_vals = _get_metric_vals(row, "mase", available)
    rmse_vals = _get_metric_vals(row, "rmse", available)

    smape_winner = min(smape_vals, key=lambda k: smape_vals[k])
    smape_best = smape_vals[smape_winner]

    ganador = smape_winner
    for mk in available:
        if mk == smape_winner:
            continue
        if smape_vals[mk] > 0 and abs(smape_vals[mk] - smape_best) / smape_vals[mk] < _EMPATE_PCT:
            mase_w = mase_vals.get(smape_winner)
            mase_c = mase_vals.get(mk)
            if mase_w is not None and mase_c is not None and mase_c < mase_w:
                ganador = mk
                break
            rmse_w = rmse_vals.get(smape_winner)
            rmse_c = rmse_vals.get(mk)
            if rmse_w is not None and rmse_c is not None and rmse_c < rmse_w:
                ganador = mk
                break

    wins = _count_wins(row, available)
    n_wins = wins.get(ganador, 0)
    total_metrics = sum(
        1 for m in _METRICS if any(pd.notna(row.get(f"{m}_{mk}")) for mk in available)
    )

    ganador_label = _MODEL_LABELS.get(ganador, ganador)
    smape_ganador = smape_vals[ganador]
    mase_ganador = mase_vals.get(ganador)
    parts: list[str] = []

    otros = sorted([(mk, v) for mk, v in smape_vals.items() if mk != ganador], key=lambda x: x[1])
    if otros:
        segundo_key, smape_segundo = otros[0]
        segundo_label = _MODEL_LABELS.get(segundo_key, segundo_key)
        if smape_ganador <= smape_segundo and smape_segundo > 0:
            pct_diff = (smape_segundo - smape_ganador) / smape_segundo * 100
            if pct_diff < 5:
                parts.append(
                    f"{ganador_label} gana por margen mínimo "
                    f"(SMAPE {smape_ganador:.1f}% vs {segundo_label} {smape_segundo:.1f}%, "
                    f"-{pct_diff:.1f}%)."
                )
            else:
                parts.append(
                    f"{ganador_label} domina con SMAPE {smape_ganador:.1f}% "
                    f"(vs {segundo_label} {smape_segundo:.1f}%)."
                )
        else:
            parts.append(
                f"{ganador_label} elegido por desempate MASE/RMSE "
                f"(SMAPE {smape_ganador:.1f}% vs {segundo_label} {smape_segundo:.1f}%)."
            )
    else:
        parts.append(f"{ganador_label} único modelo disponible (SMAPE {smape_ganador:.1f}%).")

    if mase_ganador is not None:
        if mase_ganador < 0.5:
            parts.append(f"MASE={mase_ganador:.2f}, muy superior al naive seasonal.")
        elif mase_ganador < 1.0:
            parts.append(f"MASE={mase_ganador:.2f}, supera naive seasonal.")
        else:
            parts.append(f"MASE={mase_ganador:.2f}, no supera naive seasonal.")
    if n_wins > 1:
        parts.append(f"Gana en {n_wins}/{total_metrics} métricas.")
    if smape_ganador > 150:
        parts.append("Serie de bajo volumen.")
    rmse_ganador = rmse_vals.get(ganador)
    if rmse_ganador is not None:
        parts.append(f"RMSE={rmse_ganador:.2f}.")

    return (ganador, n_wins, " ".join(parts))


# ---------------------------------------------------------------------------
# Funciones auxiliares de datos
# ---------------------------------------------------------------------------


def _display_entidad(row: pd.Series) -> str:  # type: ignore[type-arg]
    ent = row.get("Entidad", "")
    if not ent or ent == "":
        return "Nacional"
    return str(ent)


def _build_region_map() -> dict[str, str]:
    real_path = Path("data") / "processed" / "data_inegi_General.csv"
    if not real_path.exists():
        return {}
    real = pd.read_csv(real_path)
    if "region_salud_mental" not in real.columns:
        return {}
    return dict(real.groupby("Entidad")["region_salud_mental"].first().items())


def _load_forecasts() -> dict[str, pd.DataFrame]:
    forecasts: dict[str, pd.DataFrame] = {}
    base = Path("reports") / "forecasts"
    for mk in _MODELS:
        csv_path = base / mk / f"all_forecast_{mk}.csv"
        if not csv_path.exists():
            logger.warning("Forecast no encontrado: {}", csv_path)
            continue
        df = pd.read_csv(csv_path)
        df["ds"] = pd.to_datetime(df["ds"], errors="coerce")
        forecasts[mk] = df
        logger.info("  Forecast cargado {}: {} filas", mk, len(df))
    return forecasts


def _load_real_data() -> pd.DataFrame:
    path = Path("data") / "processed" / "data_inegi_General.csv"
    if not path.exists():
        logger.warning("No se encontró datos reales: {}", path)
        return pd.DataFrame()
    df = pd.read_csv(path)
    df["Fecha"] = pd.to_datetime(df["Fecha"], errors="coerce")
    return df


def _normalize_entidad(entidad: str) -> str:
    """Convierte region_X a Region X para match con forecast."""
    if entidad.startswith("region_"):
        return "Region " + entidad[len("region_") :]
    return entidad


def _sum_forecast_52(
    forecasts: dict[str, pd.DataFrame],
    modelo_key: str,
    padecimiento: str,
    entidad: str,
    sexo: str,
) -> float:
    if modelo_key not in forecasts:
        return np.nan
    df = forecasts[modelo_key]
    modo = _SEXO_TO_MODO.get(sexo, sexo)
    meta_ent = _normalize_entidad(entidad)
    mask = (
        (df["meta_padecimiento"] == padecimiento)
        & (df["meta_entidad"] == meta_ent)
        & (df["meta_modo"] == modo)
    )
    serie = df.loc[mask].sort_values("ds")
    if serie.empty:
        return np.nan
    last_52 = serie.tail(_HORIZON)
    total = last_52["yhat"].sum()
    return int(round(max(total, 0.0)))


# ---------------------------------------------------------------------------
# Diagnosticos
# ---------------------------------------------------------------------------


def _overfitting_label(smape_test: float | None, smape_train: float | None) -> str:
    if (
        smape_test is None
        or smape_train is None
        or (isinstance(smape_test, float) and np.isnan(smape_test))
        or (isinstance(smape_train, float) and np.isnan(smape_train))
        or smape_train == 0
    ):
        return "N/D"
    ratio = smape_test / smape_train
    if ratio > _OVERFIT_ALTO:
        return f"Alto ({ratio:.1f}x)"
    if ratio > _OVERFIT_MODERADO:
        return f"Moderado ({ratio:.1f}x)"
    return f"OK ({ratio:.1f}x)"


def _leakage_label(smape_train: float | None) -> str:
    if smape_train is None or (isinstance(smape_train, float) and np.isnan(smape_train)):
        return "N/D"
    if smape_train < _LEAKAGE_THRESHOLD:
        return f"Sospechoso ({smape_train:.2f}%)"
    return f"OK ({smape_train:.1f}%)"


# ---------------------------------------------------------------------------
# Series semanales (52 semanas previas)
# ---------------------------------------------------------------------------


def _get_real_series(
    real_df: pd.DataFrame, padecimiento: str, entidad: str, sexo: str
) -> pd.Series:  # type: ignore[type-arg]
    """Retorna serie de las ultimas 52 semanas reales (indexada por Fecha)."""
    if real_df.empty:
        return pd.Series(dtype=float)
    col_inc = _MODO_TO_INCREMENTO.get(sexo, "incrementos_total")
    if col_inc not in real_df.columns:
        return pd.Series(dtype=float)

    meta_ent = _normalize_entidad(entidad)

    if entidad == "Nacional":
        mask_n = real_df["Padecimiento"] == padecimiento
        agg = real_df.loc[mask_n].groupby("Fecha")[col_inc].sum().reset_index()
        agg = agg.sort_values("Fecha")
    elif entidad.startswith("region_"):
        region_name = entidad[len("region_") :]
        if "region_salud_mental" not in real_df.columns:
            return pd.Series(dtype=float)
        mask_r = (real_df["Padecimiento"] == padecimiento) & (
            real_df["region_salud_mental"] == region_name
        )
        agg = real_df.loc[mask_r].groupby("Fecha")[col_inc].sum().reset_index()
        agg = agg.sort_values("Fecha")
    else:
        mask = (real_df["Padecimiento"] == padecimiento) & (real_df["Entidad"] == meta_ent)
        agg = real_df.loc[mask].sort_values("Fecha").rename(columns={col_inc: col_inc})
    if len(agg) < _HORIZON:
        return pd.Series(dtype=float)
    last_52 = agg.tail(_HORIZON)
    return last_52.set_index("Fecha")[col_inc]


def _get_forecast_series(
    forecasts: dict[str, pd.DataFrame],
    modelo_key: str,
    padecimiento: str,
    entidad: str,
    sexo: str,
    dates: pd.Index | None = None,
) -> pd.Series:  # type: ignore[type-arg]
    """Retorna serie de yhat del modelo para las fechas dadas (52 semanas)."""
    if modelo_key not in forecasts:
        return pd.Series(dtype=float)
    df = forecasts[modelo_key]
    modo = _SEXO_TO_MODO.get(sexo, sexo)
    meta_ent = _normalize_entidad(entidad)
    mask = (
        (df["meta_padecimiento"] == padecimiento)
        & (df["meta_entidad"] == meta_ent)
        & (df["meta_modo"] == modo)
    )
    serie = df.loc[mask].sort_values("ds")
    if serie.empty:
        return pd.Series(dtype=float)
    if dates is not None and len(dates) > 0:
        date_min = pd.Timestamp(dates.min())
        date_max = pd.Timestamp(dates.max())
        in_range = serie[(serie["ds"] >= date_min) & (serie["ds"] <= date_max)]
        if len(in_range) >= _HORIZON - 2:
            return in_range.set_index("ds")["yhat"]
    # Fallback: 52 semanas antes de las ultimas 52
    n = len(serie)
    if n < _HORIZON * 2:
        return pd.Series(dtype=float)
    prev_block = serie.iloc[-(2 * _HORIZON) : -_HORIZON]
    return prev_block.set_index("ds")["yhat"]


def _is_zero_row(row: pd.Series, model_keys: list[str]) -> bool:  # type: ignore[type-arg]
    for mk in model_keys:
        v = row.get(f"rmse_{mk}")
        if pd.notna(v) and float(v) > _ZERO_THRESHOLD:
            return False
    return True


# ---------------------------------------------------------------------------
# Construccion de la tabla principal (Hoja 1)
# ---------------------------------------------------------------------------


def _build_production_table(
    merged: pd.DataFrame,
    model_keys: list[str],
    region_map: dict[str, str],
    forecasts: dict[str, pd.DataFrame],
    real_df: pd.DataFrame,
) -> pd.DataFrame:
    """Construye DataFrame de 333 filas con todas las columnas."""
    rows_out: list[dict[str, object]] = []

    for idx, (_, row) in enumerate(
        merged.sort_values(["padecimiento", "Entidad", "sexo"]).iterrows(), start=1
    ):
        ganador_key, n_wins, justificacion = _select_production(row, model_keys)
        total_metrics = sum(
            1 for m in _METRICS if any(pd.notna(row.get(f"{m}_{mk}")) for mk in model_keys)
        )

        out: dict[str, object] = {
            "numero": idx,
            "padecimiento": row.get("padecimiento", ""),
            "entidad": _display_entidad(row),
            "sexo": _SEXO_DISPLAY.get(str(row.get("sexo", "")), str(row.get("sexo", ""))),
        }

        # Metricas por modelo
        for mk in _MODELS:
            for metric in _METRICS:
                val = row.get(f"{metric}_{mk}")
                out[f"{mk}_{metric}"] = round(float(val), 4) if pd.notna(val) else np.nan

        # Mejor por metrica
        for metric in _METRICS:
            out[f"mejor_{metric}"] = _best_model_for_metric(row, metric, model_keys)

        out["victorias"] = f"{n_wins}/{total_metrics}" if total_metrics > 0 else "0/0"

        # Casos proyectados 52 semanas
        modelo_fc_key = _MODEL_KEY_MAP.get(
            _MODEL_LABELS.get(ganador_key, ganador_key), ganador_key
        )
        entidad = str(out["entidad"])
        out["casos_52_semanas_futuro"] = _sum_forecast_52(
            forecasts, modelo_fc_key, str(out["padecimiento"]), entidad, str(out["sexo"])
        )

        # Metricas del modelo de produccion
        smape_prod = row.get(f"smape_{ganador_key}") if ganador_key else None
        mase_prod = row.get(f"mase_{ganador_key}") if ganador_key else None
        rmse_prod = row.get(f"rmse_{ganador_key}") if ganador_key else None
        mae_prod = row.get(f"mae_{ganador_key}") if ganador_key else None
        smape_train_prod = row.get(f"smape_train_{ganador_key}") if ganador_key else None

        out["smape_prod"] = round(float(smape_prod), 4) if pd.notna(smape_prod) else np.nan
        out["mase_prod"] = round(float(mase_prod), 4) if pd.notna(mase_prod) else np.nan
        out["rmse_prod"] = round(float(rmse_prod), 4) if pd.notna(rmse_prod) else np.nan
        out["mae_prod"] = round(float(mae_prod), 4) if pd.notna(mae_prod) else np.nan

        # Diagnosticos
        _st = float(smape_prod) if pd.notna(smape_prod) else None
        _str = float(smape_train_prod) if pd.notna(smape_train_prod) else None
        out["overfitting"] = _overfitting_label(_st, _str)
        out["leakage"] = _leakage_label(_str)

        # Historico 52 semanas
        real_series = _get_real_series(
            real_df, str(out["padecimiento"]), entidad, str(out["sexo"])
        )
        real_sum = int(round(max(real_series.sum(), 0.0))) if len(real_series) > 0 else np.nan
        out["casos_prev_52_semanas_real"] = real_sum

        # Para DeepAR usar backtest real; para otros usar forecast CSV
        if modelo_fc_key == "deepar" and _deepar_backtest:
            fc_series = _get_deepar_backtest(str(out["padecimiento"]), entidad, str(out["sexo"]))
        else:
            fc_series = _get_forecast_series(
                forecasts,
                modelo_fc_key,
                str(out["padecimiento"]),
                entidad,
                str(out["sexo"]),
                dates=real_series.index if len(real_series) > 0 else None,
            )
        fc_sum = int(round(max(fc_series.sum(), 0.0))) if len(fc_series) > 0 else np.nan
        out["casos_prev_52_semanas_pronos"] = fc_sum

        # Precision historica (pronos / real)
        if pd.notna(real_sum) and pd.notna(fc_sum) and real_sum > 0:
            out["precision_historica"] = f"{fc_sum / real_sum * 100:.1f}%"
        else:
            out["precision_historica"] = "0%"

        # Semana previa: ultimo dato real y su pronostico
        if len(real_series) > 0:
            out["realidad_sem_previa"] = int(round(max(float(real_series.iloc[-1]), 0.0)))
        else:
            out["realidad_sem_previa"] = np.nan
        if len(fc_series) > 0 and len(real_series) > 0:
            last_date = real_series.index[-1]
            closest = fc_series.index.get_indexer([last_date], method="nearest")
            if closest[0] >= 0:
                out["pron_sem_previa"] = int(round(max(float(fc_series.iloc[closest[0]]), 0.0)))
            else:
                out["pron_sem_previa"] = int(round(max(float(fc_series.iloc[-1]), 0.0)))
        else:
            out["pron_sem_previa"] = np.nan

        # Tipo de modelo
        if _is_zero_row(row, model_keys) and entidad in region_map:
            region = region_map[entidad]
            out["tipo_modelo"] = "regional"
            out["region_asignada"] = f"region_{region}"
        else:
            out["tipo_modelo"] = "propio"
            out["region_asignada"] = "n/a"

        out["modelo_produccion"] = _MODEL_LABELS.get(ganador_key, ganador_key)
        out["justificacion"] = justificacion
        # Guardar el key del modelo para la hoja 2
        out["_ganador_key"] = ganador_key
        out["_modelo_fc_key"] = modelo_fc_key

        rows_out.append(out)

    return pd.DataFrame(rows_out)


# ---------------------------------------------------------------------------
# Post-procesamiento: reasignaciones regionales
# ---------------------------------------------------------------------------


def _apply_regional_reassignments(
    df_out: pd.DataFrame, region_map: dict[str, str]
) -> pd.DataFrame:
    """Reasigna modelo regional a filas sin incidencia o con baja confianza."""
    # 1. Sin incidencia
    zero_mask = df_out["tipo_modelo"] == "regional"
    if zero_mask.any():
        for idx_z in df_out[zero_mask].index:
            pad = df_out.at[idx_z, "padecimiento"]
            sexo = df_out.at[idx_z, "sexo"]
            region = df_out.at[idx_z, "region_asignada"]
            region_row = df_out[
                (df_out["entidad"] == region)
                & (df_out["padecimiento"] == pad)
                & (df_out["sexo"] == sexo)
            ]
            if not region_row.empty:
                modelo_regional = region_row.iloc[0]["modelo_produccion"]
                casos_regional = region_row.iloc[0]["casos_52_semanas_futuro"]
                df_out.at[idx_z, "modelo_produccion"] = modelo_regional
                df_out.at[idx_z, "casos_52_semanas_futuro"] = casos_regional
                df_out.at[idx_z, "justificacion"] = (
                    f"Sin incidencia local. Se asigna modelo de la región "
                    f"({region}): {modelo_regional}."
                )
                logger.info(
                    "  {} {} {}: sin incidencia -> {} ({})",
                    pad,
                    df_out.at[idx_z, "entidad"],
                    sexo,
                    modelo_regional,
                    region,
                )

    # 2. Baja confianza
    _no_reasignable = {"Nacional"} | {
        e for e in df_out["entidad"].unique() if str(e).startswith("region_")
    }
    low_mask = (
        (df_out["casos_52_semanas_futuro"] < _LOW_VOLUME_THRESHOLD)
        & (df_out["tipo_modelo"] == "propio")
        & (~df_out["entidad"].isin(_no_reasignable))
    )
    if low_mask.any():
        logger.info(
            "--- Reasignación por baja confianza (<{} casos/52sem) ---", _LOW_VOLUME_THRESHOLD
        )
        for idx_l in df_out[low_mask].index:
            ent = df_out.at[idx_l, "entidad"]
            if ent not in region_map:
                continue
            pad = df_out.at[idx_l, "padecimiento"]
            sexo = df_out.at[idx_l, "sexo"]
            casos_orig = df_out.at[idx_l, "casos_52_semanas_futuro"]
            modelo_orig = df_out.at[idx_l, "modelo_produccion"]
            region = region_map[ent]
            region_key = f"region_{region}"
            region_row = df_out[
                (df_out["entidad"] == region_key)
                & (df_out["padecimiento"] == pad)
                & (df_out["sexo"] == sexo)
            ]
            if not region_row.empty:
                modelo_regional = region_row.iloc[0]["modelo_produccion"]
                casos_regional = region_row.iloc[0]["casos_52_semanas_futuro"]
                df_out.at[idx_l, "modelo_produccion"] = modelo_regional
                df_out.at[idx_l, "casos_52_semanas_futuro"] = casos_regional
                df_out.at[idx_l, "tipo_modelo"] = "regional"
                df_out.at[idx_l, "region_asignada"] = region_key
                df_out.at[idx_l, "justificacion"] = (
                    f"Baja confianza: {casos_orig} casos proyectados en 52 semanas "
                    f"(modelo local: {modelo_orig}). "
                    f"Se asigna modelo de la región ({region_key}): {modelo_regional}."
                )
                logger.info(
                    "  {} {} {}: {} casos -> {} ({})",
                    pad,
                    ent,
                    sexo,
                    casos_orig,
                    modelo_regional,
                    region_key,
                )

    return df_out


# ---------------------------------------------------------------------------
# Construccion de la hoja de detalle semanal (Hoja 2)
# ---------------------------------------------------------------------------


def _build_weekly_detail(
    df_prod: pd.DataFrame,
    forecasts: dict[str, pd.DataFrame],
    real_df: pd.DataFrame,
) -> pd.DataFrame:
    """Construye DataFrame con 52 semanas de realidad, pronostico y % acierto."""
    rows: list[dict[str, object]] = []

    for _, row in df_prod.iterrows():
        pad = str(row["padecimiento"])
        ent = str(row["entidad"])
        sexo = str(row["sexo"])
        modelo_fc_key = str(row.get("_modelo_fc_key", ""))

        base: dict[str, object] = {
            "numero": row["numero"],
            "padecimiento": pad,
            "entidad": ent,
            "sexo": sexo,
            "modelo_produccion": row["modelo_produccion"],
            "tipo_modelo": row["tipo_modelo"],
            "region_asignada": row["region_asignada"],
        }

        real_series = _get_real_series(real_df, pad, ent, sexo)
        # Para DeepAR usar backtest real; para otros usar forecast CSV
        if modelo_fc_key == "deepar" and _deepar_backtest:
            fc_series = _get_deepar_backtest(pad, ent, sexo)
        else:
            fc_series = _get_forecast_series(
                forecasts,
                modelo_fc_key,
                pad,
                ent,
                sexo,
                dates=real_series.index if len(real_series) > 0 else None,
            )

        # Rellenar las 52 columnas
        for i in range(1, _HORIZON + 1):
            # Real
            if len(real_series) >= i:
                val_r = float(real_series.iloc[i - 1])
                base[f"real_sem_{i}"] = int(round(max(val_r, 0.0)))
            else:
                base[f"real_sem_{i}"] = np.nan

            # Pronostico
            if len(fc_series) >= i:
                val_p = float(fc_series.iloc[i - 1])
                base[f"pron_sem_{i}"] = int(round(max(val_p, 0.0)))
            else:
                base[f"pron_sem_{i}"] = np.nan

            # % acierto
            r_val = base[f"real_sem_{i}"]
            p_val = base[f"pron_sem_{i}"]
            if (
                pd.notna(r_val)
                and pd.notna(p_val)
                and isinstance(r_val, int | float)
                and isinstance(p_val, int | float)
                and r_val > 0
            ):
                base[f"acierto_sem_{i}"] = f"{p_val / r_val * 100:.0f}%"
            else:
                base[f"acierto_sem_{i}"] = "0%"

        rows.append(base)

    return pd.DataFrame(rows)


# ---------------------------------------------------------------------------
# Escritura Excel con formato
# ---------------------------------------------------------------------------


def _write_excel(df_prod: pd.DataFrame, df_detail: pd.DataFrame, output: Path) -> None:
    """Escribe ambas hojas en un workbook con formato IMSS."""
    from openpyxl import Workbook
    from scripts.excel_produccion_fmt import format_detail_header, format_sheet

    wb = Workbook()

    # --- Hoja 1: Produccion ---
    ws1 = wb.active
    ws1.title = "Produccion"

    # Orden final de columnas para Hoja 1
    col_order = [
        "numero",
        "padecimiento",
        "entidad",
        "sexo",
        # Metricas por modelo
        *[f"{mk}_{m}" for mk in _MODELS for m in _METRICS],
        # Mejor por metrica
        *[f"mejor_{m}" for m in _METRICS],
        "victorias",
        # Proyeccion y diagnostico
        "casos_52_semanas_futuro",
        "smape_prod",
        "mase_prod",
        "rmse_prod",
        "mae_prod",
        "overfitting",
        "leakage",
        "casos_prev_52_semanas_real",
        "casos_prev_52_semanas_pronos",
        "precision_historica",
        "pron_sem_previa",
        "realidad_sem_previa",
        # Seleccion
        "modelo_produccion",
        "tipo_modelo",
        "region_asignada",
        "justificacion",
    ]
    df_write = df_prod[[c for c in col_order if c in df_prod.columns]]

    for r_idx, row_data in enumerate(dataframe_to_rows(df_write, index=False, header=True), 1):
        for c_idx, value in enumerate(row_data, 1):
            ws1.cell(row=r_idx, column=c_idx, value=value)

    format_sheet(ws1, freeze_col=4)

    # --- Hoja 2: Detalle Semanal ---
    ws2 = wb.create_sheet("Detalle Semanal")

    for r_idx, row_data in enumerate(dataframe_to_rows(df_detail, index=False, header=True), 1):
        for c_idx, value in enumerate(row_data, 1):
            ws2.cell(row=r_idx, column=c_idx, value=value)

    format_sheet(ws2, freeze_col=5)
    format_detail_header(ws2)

    # --- Hoja 3: Analisis Visual ---
    logger.info("Generando graficos para hoja Análisis Visual...")
    from openpyxl.drawing.image import Image as XlImage
    from scripts.excel_produccion_charts import generate_all_charts

    ws3 = wb.create_sheet("Análisis Visual")
    ws3.sheet_properties.tabColor = "006847"

    charts = generate_all_charts(df_prod)
    row_cursor = 1
    for name, buf in charts:
        img = XlImage(buf)
        img.width = int(img.width * 0.75)
        img.height = int(img.height * 0.75)
        ws3.add_image(img, f"A{row_cursor}")
        # Estimar filas que ocupa la imagen (~20px por fila)
        rows_needed = max(int(img.height / 20) + 2, 20)
        row_cursor += rows_needed
        logger.info("  Grafico embebido: {}", name)

    # Guardar
    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output))


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------


def main() -> None:
    """Genera el Excel de 333 modelos de produccion."""
    logger.info("=== Tabla de modelos de producción ===")

    # 1. Cargar y merge
    data = cargar_completos()
    if not data:
        logger.error("No se encontraron CSVs. Ejecute 'make train-all' primero.")
        return

    model_keys = [mk for mk in _MODELS if mk in data]
    merged = merge_all_models(data)
    logger.info("Merge: {} filas, modelos: {}", len(merged), model_keys)

    region_map = _build_region_map()
    forecasts = _load_forecasts()
    real_df = _load_real_data()

    # 1b. Backtest DeepAR (fitted values reales, no copias del historico)
    global _deepar_backtest  # noqa: PLW0603
    _deepar_backtest = _build_deepar_backtest_cache()

    # 2. Construir tabla principal
    df_out = _build_production_table(merged, model_keys, region_map, forecasts, real_df)

    # 3. Reasignaciones regionales
    df_out = _apply_regional_reassignments(df_out, region_map)

    # 4. Construir detalle semanal (Hoja 2)
    logger.info("Construyendo detalle semanal (52 semanas x 333 series)...")
    df_detail = _build_weekly_detail(df_out, forecasts, real_df)

    # 5. Limpiar columnas internas
    internal_cols = [c for c in df_out.columns if c.startswith("_")]
    df_out = df_out.drop(columns=internal_cols)
    if "_modelo_fc_key" in df_detail.columns:
        df_detail = df_detail.drop(columns=["_modelo_fc_key"])

    # 6. Escribir Excel
    logger.info("Escribiendo Excel con formato IMSS...")
    _write_excel(df_out, df_detail, _OUTPUT)
    logger.success("Excel generado: {} ({} filas)", _OUTPUT, len(df_out))

    # 7. Resumen
    logger.info("--- Distribución de modelos de producción ---")
    counts = df_out["modelo_produccion"].value_counts()
    total = len(df_out)
    for modelo, n in counts.items():
        logger.info("  {}: {}/{} ({:.1f}%)", modelo, n, total, n / total * 100)
    for pad in sorted(df_out["padecimiento"].unique()):
        sub = df_out[df_out["padecimiento"] == pad]
        pad_counts = sub["modelo_produccion"].value_counts()
        parts = [f"{m}: {c}" for m, c in pad_counts.items()]
        logger.info("  {} -> {}", pad, ", ".join(parts))


if __name__ == "__main__":
    main()
